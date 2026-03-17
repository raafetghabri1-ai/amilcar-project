"""
AMILCAR — Reports, KPIs & Dashboards
Blueprint: reports_bp
Routes: 28
"""
from flask import Blueprint, render_template, request, redirect, url_for, flash, make_response, jsonify, session, send_file
from helpers import login_required, admin_required, client_required, get_db, get_services, get_setting, get_all_settings
from helpers import allowed_file, safe_page, log_activity, build_wa_url, STATUS_MESSAGES, UPLOAD_FOLDER, MAX_FILE_SIZE, MAX_FILES, PER_PAGE
from database.db import get_db
from werkzeug.utils import secure_filename
from werkzeug.security import generate_password_hash, check_password_hash
from datetime import datetime, date, timedelta
import os, re, uuid, io
import time as time_module
import sqlite3

reports_bp = Blueprint("reports_bp", __name__)


def generate_pdf(html_content):
    """Generate PDF from HTML using WeasyPrint."""
    from weasyprint import HTML
    pdf = HTML(string=html_content).write_pdf()
    response = make_response(pdf)
    response.headers['Content-Type'] = 'application/pdf'
    return response


@reports_bp.route("/report_pdf/<report_type>")
@login_required
def report_pdf(report_type):
    """Generate PDF for any report page."""
    allowed = ['daily', 'reports', 'end_of_day', 'advanced_report', 'ceo_dashboard']
    if report_type not in allowed:
        flash("Type de rapport invalide", "error")
        return redirect('/')
    from flask import current_app
    client = current_app.test_client()
    # Copy session to internal request
    with client.session_transaction() as sess:
        sess['user_id'] = session.get('user_id')
        sess['username'] = session.get('username')
        sess['role'] = session.get('role')
    resp = client.get(f'/{report_type}')
    if resp.status_code != 200:
        flash("Erreur lors de la génération du PDF", "error")
        return redirect(f'/{report_type}')
    html = resp.data.decode('utf-8')
    # Add print-friendly CSS
    html = html.replace('</head>', '<style>@page{size:A4;margin:1cm} body{background:#fff!important;color:#000!important} .no-print,.sidebar,.navbar{display:none!important}</style></head>')
    response = generate_pdf(html)
    response.headers['Content-Disposition'] = f'attachment; filename=rapport_{report_type}_{datetime.now().strftime("%Y%m%d")}.pdf'
    return response


@reports_bp.route("/daily")
@login_required
def daily():
    from datetime import date
    today = str(date.today())
    with get_db() as conn:
        appointments = conn.execute("SELECT a.id, cu.name, ca.brand, ca.model, a.date, a.service, a.status, COALESCE(a.time, '') FROM appointments a JOIN cars ca ON a.car_id = ca.id JOIN customers cu ON ca.customer_id = cu.id WHERE a.date = ?", (today,)).fetchall()
        revenue = conn.execute("SELECT SUM(amount) FROM invoices i JOIN appointments a ON i.appointment_id = a.id WHERE a.date = ? AND i.status = 'paid'", (today,)).fetchone()[0] or 0
        expenses_today = conn.execute("SELECT COALESCE(SUM(amount),0) FROM expenses WHERE date = ?", (today,)).fetchone()[0]
    return render_template("daily.html", appointments=appointments, revenue=revenue, today=today, expenses_today=expenses_today)



@reports_bp.route("/monthly")
@login_required
def monthly():
    from datetime import date, timedelta
    month_param = request.args.get("month")
    if month_param:
        year, mon = map(int, month_param.split("-"))
    else:
        today = date.today()
        year, mon = today.year, today.month
    month_start = f"{year}-{mon:02d}-01"
    if mon == 12:
        next_y, next_m = year + 1, 1
    else:
        next_y, next_m = year, mon + 1
    month_end = f"{next_y}-{next_m:02d}-01"
    if mon == 1:
        prev_y, prev_m = year - 1, 12
    else:
        prev_y, prev_m = year, mon - 1
    months_ar = ["JAN","FEB","MAR","APR","MAY","JUN","JUL","AUG","SEP","OCT","NOV","DEC"]
    month_label = f"{months_ar[mon-1]} {year}"
    prev_label = f"{months_ar[prev_m-1]} {prev_y}"
    next_label = f"{months_ar[next_m-1]} {next_y}"
    with get_db() as conn:
        appointments = conn.execute(
            "SELECT a.id, cu.name, ca.brand, ca.model, a.date, a.service, a.status "
            "FROM appointments a JOIN cars ca ON a.car_id = ca.id "
            "JOIN customers cu ON ca.customer_id = cu.id "
            "WHERE a.date >= ? AND a.date < ? ORDER BY a.date",
            (month_start, month_end)).fetchall()
        invoices = conn.execute(
            "SELECT i.id, cu.name, a.service, i.amount, i.status "
            "FROM invoices i JOIN appointments a ON i.appointment_id = a.id "
            "JOIN cars ca ON a.car_id = ca.id JOIN customers cu ON ca.customer_id = cu.id "
            "WHERE a.date >= ? AND a.date < ? ORDER BY i.id",
            (month_start, month_end)).fetchall()
        revenue = conn.execute(
            "SELECT COALESCE(SUM(i.amount),0) FROM invoices i JOIN appointments a ON i.appointment_id = a.id "
            "WHERE a.date >= ? AND a.date < ? AND i.status = 'paid'",
            (month_start, month_end)).fetchone()[0]
        unpaid = conn.execute(
            "SELECT COALESCE(SUM(i.amount),0) FROM invoices i JOIN appointments a ON i.appointment_id = a.id "
            "WHERE a.date >= ? AND a.date < ? AND i.status = 'unpaid'",
            (month_start, month_end)).fetchone()[0]
        completed = sum(1 for a in appointments if a[6] == 'completed')
        month_expenses = conn.execute(
            "SELECT COALESCE(SUM(amount),0) FROM expenses WHERE date >= ? AND date < ?",
            (month_start, month_end)).fetchone()[0]
    stats = {
        'appointments': len(appointments),
        'completed': completed,
        'revenue': revenue,
        'unpaid': unpaid,
        'expenses': month_expenses,
        'profit': revenue - month_expenses
    }
    return render_template("monthly.html",
        stats=stats, appointments=appointments, invoices=invoices,
        month_label=month_label,
        current_month=f"{year}-{mon:02d}",
        prev_month=f"{prev_y}-{prev_m:02d}", prev_label=prev_label,
        next_month=f"{next_y}-{next_m:02d}", next_label=next_label)



# ─── KPI Dashboard ───
@reports_bp.route("/kpi")
@login_required
def kpi_dashboard():
    from datetime import date, timedelta
    with get_db() as conn:
        today = date.today()
        # Current month boundaries
        ms = f"{today.year}-{today.month:02d}-01"
        if today.month == 12:
            me = f"{today.year+1}-01-01"
        else:
            me = f"{today.year}-{today.month+1:02d}-01"
        # Previous month boundaries
        if today.month == 1:
            pms = f"{today.year-1}-12-01"
            pme = ms
        else:
            pms = f"{today.year}-{today.month-1:02d}-01"
            pme = ms
        # Current month stats
        curr_revenue = conn.execute(
            "SELECT COALESCE(SUM(i.amount),0) FROM invoices i JOIN appointments a ON i.appointment_id = a.id "
            "WHERE a.date >= ? AND a.date < ? AND i.status = 'paid'", (ms, me)).fetchone()[0]
        curr_appts = conn.execute("SELECT COUNT(*) FROM appointments WHERE date >= ? AND date < ?", (ms, me)).fetchone()[0]
        curr_completed = conn.execute("SELECT COUNT(*) FROM appointments WHERE date >= ? AND date < ? AND status = 'completed'", (ms, me)).fetchone()[0]
        curr_new_customers = conn.execute(
            "SELECT COUNT(DISTINCT ca.customer_id) FROM appointments a JOIN cars ca ON a.car_id = ca.id "
            "WHERE a.date >= ? AND a.date < ? AND ca.customer_id NOT IN "
            "(SELECT DISTINCT ca2.customer_id FROM appointments a2 JOIN cars ca2 ON a2.car_id = ca2.id WHERE a2.date < ?)",
            (ms, me, ms)).fetchone()[0]
        # Previous month stats
        prev_revenue = conn.execute(
            "SELECT COALESCE(SUM(i.amount),0) FROM invoices i JOIN appointments a ON i.appointment_id = a.id "
            "WHERE a.date >= ? AND a.date < ? AND i.status = 'paid'", (pms, pme)).fetchone()[0]
        prev_appts = conn.execute("SELECT COUNT(*) FROM appointments WHERE date >= ? AND date < ?", (pms, pme)).fetchone()[0]
        # Completion rate
        completion_rate = round(curr_completed / curr_appts * 100) if curr_appts > 0 else 0
        # Average revenue per day (this month)
        days_elapsed = max(1, (today - date(today.year, today.month, 1)).days + 1)
        avg_daily = round(curr_revenue / days_elapsed, 1)
        # Average rating this month
        avg_rating = conn.execute(
            "SELECT AVG(r.rating), COUNT(*) FROM ratings r WHERE r.created_at >= ?", (ms,)).fetchone()
        # Returning customers rate
        total_customers_visited = conn.execute(
            "SELECT COUNT(DISTINCT ca.customer_id) FROM appointments a JOIN cars ca ON a.car_id = ca.id "
            "WHERE a.date >= ? AND a.date < ?", (ms, me)).fetchone()[0]
        returning = total_customers_visited - curr_new_customers if total_customers_visited > curr_new_customers else 0
        return_rate = round(returning / total_customers_visited * 100) if total_customers_visited > 0 else 0
        # Revenue growth
        revenue_growth = round((curr_revenue - prev_revenue) / prev_revenue * 100) if prev_revenue > 0 else 0
        # Top technicians this month
        top_techs = conn.execute(
            "SELECT a.assigned_to, COUNT(*) as cnt, COALESCE(SUM(i.amount),0) "
            "FROM appointments a LEFT JOIN invoices i ON i.appointment_id = a.id AND i.status = 'paid' "
            "WHERE a.date >= ? AND a.date < ? AND a.assigned_to != '' "
            "GROUP BY a.assigned_to ORDER BY cnt DESC LIMIT 5", (ms, me)).fetchall()
    kpi = {
        'revenue': curr_revenue, 'prev_revenue': prev_revenue, 'revenue_growth': revenue_growth,
        'appointments': curr_appts, 'prev_appointments': prev_appts,
        'completion_rate': completion_rate, 'avg_daily': avg_daily,
        'avg_rating': round(avg_rating[0], 1) if avg_rating[0] else 0,
        'rating_count': avg_rating[1],
        'new_customers': curr_new_customers, 'return_rate': return_rate,
        'top_techs': top_techs
    }
    month_names = ["Janvier","Février","Mars","Avril","Mai","Juin","Juillet","Août","Septembre","Octobre","Novembre","Décembre"]
    return render_template("kpi.html", kpi=kpi, month_label=f"{month_names[today.month-1]} {today.year}")



# ─── Advanced Financial Reports ───
@reports_bp.route("/reports")
@login_required
def reports():
    from datetime import date
    with get_db() as conn:
        today = date.today()
        # Monthly comparison - last 6 months
        months_data = []
        for i in range(5, -1, -1):
            m = today.month - i
            y = today.year
            while m <= 0:
                m += 12
                y -= 1
            ms = f"{y}-{m:02d}-01"
            if m == 12:
                me = f"{y+1}-01-01"
            else:
                me = f"{y}-{m+1:02d}-01"
            rev = conn.execute(
                "SELECT COALESCE(SUM(i.amount),0) FROM invoices i JOIN appointments a ON i.appointment_id = a.id "
                "WHERE a.date >= ? AND a.date < ? AND i.status = 'paid'", (ms, me)).fetchone()[0]
            exp = conn.execute(
                "SELECT COALESCE(SUM(amount),0) FROM expenses WHERE date >= ? AND date < ?", (ms, me)).fetchone()[0]
            appts = conn.execute(
                "SELECT COUNT(*) FROM appointments WHERE date >= ? AND date < ?", (ms, me)).fetchone()[0]
            month_names = ["JAN","FEB","MAR","APR","MAY","JUN","JUL","AUG","SEP","OCT","NOV","DEC"]
            months_data.append({
                'label': f"{month_names[m-1]} {y}", 'revenue': rev,
                'expenses': exp, 'profit': rev - exp, 'appointments': appts
            })
        # Top 5 most profitable services
        top_services = conn.execute(
            "SELECT a.service, COUNT(*) as cnt, COALESCE(SUM(i.amount),0) as total "
            "FROM appointments a LEFT JOIN invoices i ON i.appointment_id = a.id AND i.status = 'paid' "
            "GROUP BY a.service ORDER BY total DESC LIMIT 5"
        ).fetchall()
        # Top 5 spending customers
        top_customers = conn.execute(
            "SELECT cu.id, cu.name, COALESCE(SUM(i.amount),0) as total, COUNT(DISTINCT a.id) as visits "
            "FROM customers cu JOIN cars ca ON ca.customer_id = cu.id "
            "JOIN appointments a ON a.car_id = ca.id "
            "LEFT JOIN invoices i ON i.appointment_id = a.id AND i.status = 'paid' "
            "GROUP BY cu.id ORDER BY total DESC LIMIT 5"
        ).fetchall()
        # Payment method breakdown
        payment_methods = conn.execute(
            "SELECT COALESCE(payment_method, 'N/A'), COUNT(*), COALESCE(SUM(amount),0) "
            "FROM invoices WHERE status = 'paid' GROUP BY payment_method ORDER BY SUM(amount) DESC"
        ).fetchall()
        # Invoice stats
        total_paid = conn.execute("SELECT COALESCE(SUM(amount),0) FROM invoices WHERE status = 'paid'").fetchone()[0]
        total_unpaid = conn.execute("SELECT COALESCE(SUM(amount),0) FROM invoices WHERE status = 'unpaid'").fetchone()[0]
        total_partial = conn.execute("SELECT COALESCE(SUM(amount - COALESCE(paid_amount,0)),0) FROM invoices WHERE status = 'partial'").fetchone()[0]
    return render_template("reports.html", months_data=months_data, top_services=top_services,
                           top_customers=top_customers, payment_methods=payment_methods,
                           total_paid=total_paid, total_unpaid=total_unpaid, total_partial=total_partial)



# ─── Feature 7: Service Profitability Analysis ───
@reports_bp.route("/profitability")
@login_required
def service_profitability():
    with get_db() as conn:
        # Get services with revenue and material costs
        services = conn.execute(
            "SELECT a.service, COUNT(*) as cnt, COALESCE(SUM(i.amount),0) as revenue "
            "FROM appointments a LEFT JOIN invoices i ON i.appointment_id = a.id AND i.status = 'paid' "
            "WHERE a.status = 'completed' GROUP BY a.service ORDER BY revenue DESC").fetchall()
        # Material cost per service from service_inventory
        cost_data = {}
        links = conn.execute(
            "SELECT si.service_name, SUM(si.quantity_used * inv.unit_price) as cost "
            "FROM service_inventory si JOIN inventory inv ON si.inventory_id = inv.id "
            "GROUP BY si.service_name").fetchall()
        for l in links:
            cost_data[l[0]] = l[1]
    results = []
    for s in services:
        service_name = s[0].split(' - ')[0].strip()
        material_cost = cost_data.get(service_name, 0) * s[1]
        profit = s[2] - material_cost
        margin = round(profit / s[2] * 100) if s[2] > 0 else 0
        results.append({
            'service': s[0], 'count': s[1], 'revenue': s[2],
            'material_cost': round(material_cost, 1),
            'profit': round(profit, 1), 'margin': margin
        })
    return render_template("profitability.html", services=results)



# ─── Phase 6 Feature 1: Advanced PDF Reports ───
@reports_bp.route("/advanced_report")
@login_required
def advanced_report():
    from datetime import date, timedelta
    period = request.args.get('period', 'month')
    today = date.today()
    if period == 'year':
        start = f"{today.year}-01-01"
        title = f"Rapport Annuel {today.year}"
    else:
        start = f"{today.year}-{today.month:02d}-01"
        title = f"Rapport Mensuel {today.strftime('%B %Y')}"
    end = today.isoformat()
    with get_db() as conn:
        revenue = conn.execute("SELECT COALESCE(SUM(amount),0) FROM invoices WHERE status='paid' AND created_at BETWEEN ? AND ?", (start, end)).fetchone()[0]
        expenses = conn.execute("SELECT COALESCE(SUM(amount),0) FROM expenses WHERE date BETWEEN ? AND ?", (start, end)).fetchone()[0]
        appt_count = conn.execute("SELECT COUNT(*) FROM appointments WHERE date BETWEEN ? AND ?", (start, end)).fetchone()[0]
        completed = conn.execute("SELECT COUNT(*) FROM appointments WHERE date BETWEEN ? AND ? AND status='completed'", (start, end)).fetchone()[0]
        new_customers = conn.execute("SELECT COUNT(*) FROM customers WHERE id IN (SELECT DISTINCT ca.customer_id FROM cars ca JOIN appointments a ON a.car_id=ca.id WHERE a.date BETWEEN ? AND ?)", (start, end)).fetchone()[0]
        top_services = conn.execute("SELECT service, COUNT(*) as cnt, COALESCE(SUM(i.amount),0) FROM appointments a LEFT JOIN invoices i ON i.appointment_id=a.id AND i.status='paid' WHERE a.date BETWEEN ? AND ? GROUP BY a.service ORDER BY cnt DESC LIMIT 10", (start, end)).fetchall()
        top_customers = conn.execute("SELECT cu.name, COUNT(*) as cnt, COALESCE(SUM(i.amount),0) FROM appointments a JOIN cars ca ON a.car_id=ca.id JOIN customers cu ON ca.customer_id=cu.id LEFT JOIN invoices i ON i.appointment_id=a.id AND i.status='paid' WHERE a.date BETWEEN ? AND ? GROUP BY cu.id ORDER BY cnt DESC LIMIT 10", (start, end)).fetchall()
        monthly_rev = conn.execute("SELECT strftime('%Y-%m', created_at) as m, SUM(amount) FROM invoices WHERE status='paid' AND created_at >= date(?, '-12 months') GROUP BY m ORDER BY m", (end,)).fetchall()
    data = {
        'title': title, 'period': period, 'start': start, 'end': end,
        'revenue': revenue, 'expenses': expenses, 'profit': revenue - expenses,
        'appt_count': appt_count, 'completed': completed, 'new_customers': new_customers,
        'completion_rate': round(completed/appt_count*100) if appt_count else 0,
        'top_services': [tuple(r) for r in top_services], 'top_customers': [tuple(r) for r in top_customers], 'monthly_rev': [tuple(r) for r in monthly_rev]
    }
    fmt = request.args.get('format', 'html')
    if fmt == 'pdf':
        from xhtml2pdf import pisa
        html = render_template("advanced_report.html", data=data, pdf_mode=True)
        result = io.BytesIO()
        pisa.CreatePDF(io.BytesIO(html.encode('utf-8')), dest=result)
        result.seek(0)
        response = make_response(result.read())
        response.headers['Content-Type'] = 'application/pdf'
        response.headers['Content-Disposition'] = f'attachment; filename=rapport_{period}_{end}.pdf'
        return response
    return render_template("advanced_report.html", data=data, pdf_mode=False)



# ─── Phase 7 Feature 1: CEO Dashboard ───
@reports_bp.route("/ceo_dashboard")
@login_required
def ceo_dashboard():
    with get_db() as conn:
        from datetime import date, timedelta
        today = date.today()
        month_start = today.replace(day=1).isoformat()
        last_month_start = (today.replace(day=1) - timedelta(days=1)).replace(day=1).isoformat()
        last_month_end = (today.replace(day=1) - timedelta(days=1)).isoformat()
        year_start = today.replace(month=1, day=1).isoformat()

        # Revenue this month
        rev_month = conn.execute("SELECT COALESCE(SUM(amount),0) FROM invoices WHERE created_at >= ? AND status='Payée'", (month_start,)).fetchone()[0]
        # Revenue last month
        rev_last = conn.execute("SELECT COALESCE(SUM(amount),0) FROM invoices WHERE created_at >= ? AND created_at <= ? AND status='Payée'", (last_month_start, last_month_end)).fetchone()[0]
        # Revenue this year
        rev_year = conn.execute("SELECT COALESCE(SUM(amount),0) FROM invoices WHERE created_at >= ? AND status='Payée'", (year_start,)).fetchone()[0]
        # Expenses this month
        exp_month = conn.execute("SELECT COALESCE(SUM(amount),0) FROM expenses WHERE date >= ?", (month_start,)).fetchone()[0]
        # Expenses last month
        exp_last = conn.execute("SELECT COALESCE(SUM(amount),0) FROM expenses WHERE date >= ? AND date <= ?", (last_month_start, last_month_end)).fetchone()[0]
        # Net profit
        profit_month = rev_month - exp_month
        profit_last = rev_last - exp_last
        # Clients total & new this month
        total_clients = conn.execute("SELECT COUNT(*) FROM customers").fetchone()[0]
        new_clients = conn.execute("SELECT COUNT(*) FROM customers WHERE id IN (SELECT DISTINCT ca.customer_id FROM cars ca JOIN appointments a ON a.car_id=ca.id WHERE a.date >= ?)", (month_start,)).fetchone()[0]
        # Appointments this month
        appts_month = conn.execute("SELECT COUNT(*) FROM appointments WHERE date >= ?", (month_start,)).fetchone()[0]
        # Average rating
        avg_rating = conn.execute("SELECT COALESCE(AVG(rating),0) FROM ratings").fetchone()[0]
        # Unpaid invoices
        unpaid_total = conn.execute("SELECT COALESCE(SUM(amount),0) FROM invoices WHERE status IN ('unpaid','Non payée','partial')").fetchone()[0]
        unpaid_count = conn.execute("SELECT COUNT(*) FROM invoices WHERE status IN ('unpaid','Non payée','partial')").fetchone()[0]
        # Monthly revenue trend (last 12 months)
        monthly_data = []
        for i in range(11, -1, -1):
            m = today.replace(day=1) - timedelta(days=i*30)
            ms = m.replace(day=1).isoformat()
            if m.month == 12:
                me = m.replace(year=m.year+1, month=1, day=1).isoformat()
            else:
                me = m.replace(month=m.month+1, day=1).isoformat()
            r = conn.execute("SELECT COALESCE(SUM(amount),0) FROM invoices WHERE created_at >= ? AND created_at < ? AND status='Payée'", (ms, me)).fetchone()[0]
            e = conn.execute("SELECT COALESCE(SUM(amount),0) FROM expenses WHERE date >= ? AND date < ?", (ms, me)).fetchone()[0]
            monthly_data.append({'month': ms[:7], 'revenue': r, 'expenses': e, 'profit': r - e})
        # Top services
        top_services = [tuple(r) for r in conn.execute("SELECT service, COUNT(*) as cnt FROM appointments WHERE date >= ? GROUP BY service ORDER BY cnt DESC LIMIT 5", (year_start,)).fetchall()]
        # Top customers by spending
        top_customers = [tuple(r) for r in conn.execute("""
            SELECT c.name, COALESCE(SUM(i.amount),0) as total FROM invoices i
            JOIN appointments a ON i.appointment_id = a.id
            JOIN cars cr ON a.car_id = cr.id
            JOIN customers c ON cr.customer_id = c.id
            WHERE i.status='Payée' GROUP BY c.id ORDER BY total DESC LIMIT 5
        """).fetchall()]

    return render_template("ceo_dashboard.html",
        rev_month=rev_month, rev_last=rev_last, rev_year=rev_year,
        exp_month=exp_month, exp_last=exp_last,
        profit_month=profit_month, profit_last=profit_last,
        total_clients=total_clients, new_clients=new_clients,
        appts_month=appts_month, avg_rating=round(avg_rating, 1),
        unpaid_total=unpaid_total, unpaid_count=unpaid_count,
        monthly_data=monthly_data, top_services=top_services,
        top_customers=top_customers, now=today.isoformat())



# ─── Phase 8 Feature 8: Retention Analysis ───
@reports_bp.route("/retention_analysis")
@login_required
def retention_analysis():
    with get_db() as conn:
        from datetime import date, timedelta
        today = date.today()
        # All customers with visits
        customers_data = conn.execute("""
            SELECT c.id, c.name, c.phone, COUNT(a.id) as visits,
                   MIN(a.date) as first_visit, MAX(a.date) as last_visit,
                   COALESCE(SUM(i.amount),0) as total_spent
            FROM customers c
            LEFT JOIN cars cr ON cr.customer_id=c.id
            LEFT JOIN appointments a ON a.car_id=cr.id
            LEFT JOIN invoices i ON i.appointment_id=a.id AND i.status='Payée'
            GROUP BY c.id ORDER BY last_visit DESC
        """).fetchall()
        total = len(customers_data)
        active_30 = len([c for c in customers_data if c[5] and c[5] >= (today - timedelta(days=30)).isoformat()])
        active_90 = len([c for c in customers_data if c[5] and c[5] >= (today - timedelta(days=90)).isoformat()])
        churned = len([c for c in customers_data if c[5] and c[5] < (today - timedelta(days=90)).isoformat()])
        new_30 = len([c for c in customers_data if c[4] and c[4] >= (today - timedelta(days=30)).isoformat()])
        returning = len([c for c in customers_data if c[3] and c[3] > 1])
        retention_rate = round(returning / total * 100, 1) if total > 0 else 0
        churn_rate = round(churned / total * 100, 1) if total > 0 else 0
        # Monthly retention
        monthly_retention = []
        for i in range(5, -1, -1):
            m = today.replace(day=1) - timedelta(days=i*30)
            ms = m.replace(day=1).isoformat()
            me = (m.replace(day=28) + timedelta(days=4)).replace(day=1).isoformat()
            active = conn.execute("SELECT COUNT(DISTINCT cr.customer_id) FROM appointments a JOIN cars cr ON a.car_id=cr.id WHERE a.date >= ? AND a.date < ?", (ms, me)).fetchone()[0]
            monthly_retention.append({'month': ms[:7], 'active': active})
        # At risk (visited 2+ times, last visit 30-90 days ago)
        at_risk = [c for c in customers_data if c[3] >= 2 and c[5] and
                   (today - timedelta(days=90)).isoformat() <= c[5] < (today - timedelta(days=30)).isoformat()]
    return render_template("retention_analysis.html",
        total=total, active_30=active_30, active_90=active_90, churned=churned,
        new_30=new_30, returning=returning, retention_rate=retention_rate,
        churn_rate=churn_rate, monthly_retention=monthly_retention,
        at_risk=at_risk[:20], customers=customers_data[:50])



# ─── 2. Export Data (Excel/CSV/PDF) ───
@reports_bp.route("/export_data")
@login_required
def export_data():
    return render_template("export_data.html")



@reports_bp.route("/export_data/<data_type>/<fmt>")
@login_required
def export_data_download(data_type, fmt):
    import csv
    allowed_types = ['customers', 'invoices', 'appointments', 'cars', 'expenses']
    allowed_fmts = ['csv', 'pdf']
    if data_type not in allowed_types or fmt not in allowed_fmts:
        flash("Type ou format non supporté", "error")
        return redirect("/export_data")

    with get_db() as conn:
        if data_type == 'customers':
            rows = conn.execute("SELECT id, name, phone, email, notes FROM customers ORDER BY name").fetchall()
            headers = ['ID', 'Nom', 'Téléphone', 'Email', 'Notes']
        elif data_type == 'invoices':
            rows = conn.execute("""SELECT i.id, c.name, ca.plate, a.service, i.amount, i.status, i.payment_method
                FROM invoices i JOIN appointments a ON i.appointment_id=a.id
                JOIN cars ca ON a.car_id=ca.id JOIN customers c ON ca.customer_id=c.id
                ORDER BY i.id DESC""").fetchall()
            headers = ['ID', 'Client', 'Plaque', 'Service', 'Montant', 'Statut', 'Paiement']
        elif data_type == 'appointments':
            rows = conn.execute("""SELECT a.id, c.name, ca.plate, a.service, a.date, a.time, a.status, a.assigned_to
                FROM appointments a JOIN cars ca ON a.car_id=ca.id JOIN customers c ON ca.customer_id=c.id
                ORDER BY a.date DESC""").fetchall()
            headers = ['ID', 'Client', 'Plaque', 'Service', 'Date', 'Heure', 'Statut', 'Technicien']
        elif data_type == 'cars':
            rows = conn.execute("""SELECT ca.id, c.name, ca.brand, ca.model, ca.plate, ca.year, ca.color, ca.mileage
                FROM cars ca JOIN customers c ON ca.customer_id=c.id ORDER BY c.name""").fetchall()
            headers = ['ID', 'Propriétaire', 'Marque', 'Modèle', 'Plaque', 'Année', 'Couleur', 'Kilométrage']
        else:  # expenses
            rows = conn.execute("SELECT id, date, category, description, amount FROM expenses ORDER BY date DESC").fetchall()
            headers = ['ID', 'Date', 'Catégorie', 'Description', 'Montant']

    if fmt == 'csv':
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(headers)
        for row in rows:
            writer.writerow(row)
        resp = make_response(output.getvalue())
        resp.headers['Content-Type'] = 'text/csv; charset=utf-8'
        resp.headers['Content-Disposition'] = f'attachment; filename=amilcar_{data_type}.csv'
        return resp
    else:  # pdf
        try:
            from xhtml2pdf import pisa
        except ImportError:
            flash("xhtml2pdf non installé", "error")
            return redirect("/export_data")
        html = f"<html><head><meta charset='utf-8'><style>body{{font-family:Helvetica,sans-serif;font-size:11px}}table{{width:100%;border-collapse:collapse}}th,td{{border:1px solid #ccc;padding:5px;text-align:left}}th{{background:#D4AF37;color:#fff}}</style></head><body>"
        html += f"<h2>AMILCAR — {data_type.upper()}</h2><table><tr>"
        for h in headers:
            html += f"<th>{h}</th>"
        html += "</tr>"
        for row in rows:
            html += "<tr>" + "".join(f"<td>{c}</td>" for c in row) + "</tr>"
        html += "</table></body></html>"
        pdf_out = io.BytesIO()
        pisa.CreatePDF(io.StringIO(html), dest=pdf_out)
        resp = make_response(pdf_out.getvalue())
        resp.headers['Content-Type'] = 'application/pdf'
        resp.headers['Content-Disposition'] = f'attachment; filename=amilcar_{data_type}.pdf'
        return resp



@reports_bp.route("/quality_check/<int:appt_id>", methods=["GET", "POST"])
@login_required
def quality_check(appt_id):
    import json
    with get_db() as conn:
        appt = conn.execute("""SELECT a.*, c.plate, cu.name as customer_name 
            FROM appointments a JOIN cars c ON a.car_id=c.id 
            JOIN customers cu ON c.customer_id=cu.id WHERE a.id=?""", (appt_id,)).fetchone()
        if not appt:
            flash("RDV introuvable", "danger")
            return redirect("/appointments")
        existing = conn.execute("SELECT * FROM quality_checks WHERE appointment_id=?", (appt_id,)).fetchone()
        if request.method == "POST":
            checklist = []
            for item in QUALITY_CHECKLIST:
                status = request.form.get(f"check_{item['id']}", "pending")
                note = request.form.get(f"note_{item['id']}", "").strip()
                checklist.append({'id': item['id'], 'label': item['label'], 'category': item['category'],
                                'status': status, 'note': note})
            passed = sum(1 for c in checklist if c['status'] == 'pass')
            total = len(checklist)
            score = int((passed / total) * 100) if total else 0
            nps = request.form.get("nps_score", 0, type=int)
            nps_comment = request.form.get("nps_comment", "").strip()
            checklist_json = json.dumps(checklist)
            if existing:
                conn.execute("""UPDATE quality_checks SET checklist=?, overall_score=?, nps_score=?, 
                    nps_comment=?, status=? WHERE appointment_id=?""",
                    (checklist_json, score, nps, nps_comment, 'completed' if score >= 80 else 'needs_review', appt_id))
            else:
                conn.execute("""INSERT INTO quality_checks 
                    (appointment_id, inspector_id, checklist, overall_score, nps_score, nps_comment, status)
                    VALUES (?,?,?,?,?,?,?)""",
                    (appt_id, session.get('user_id', 0), checklist_json, score, nps, nps_comment,
                     'completed' if score >= 80 else 'needs_review'))
            conn.commit()
            flash(f"Contrôle qualité enregistré — Score: {score}%", "success")
            return redirect(f"/quality_check/{appt_id}")
        parsed_checklist = json.loads(existing['checklist']) if existing and existing['checklist'] else []
    return render_template("quality_check.html", appt=appt, existing=existing,
                          checklist=QUALITY_CHECKLIST, parsed=parsed_checklist)



@reports_bp.route("/quality_dashboard")
@login_required
def quality_dashboard():
    import json
    with get_db() as conn:
        checks = conn.execute("""SELECT qc.*, a.service, a.date, c.plate, cu.name as customer_name
            FROM quality_checks qc 
            JOIN appointments a ON qc.appointment_id=a.id 
            JOIN cars c ON a.car_id=c.id 
            JOIN customers cu ON c.customer_id=cu.id 
            ORDER BY qc.created_at DESC LIMIT 100""").fetchall()
        avg_score = conn.execute("SELECT AVG(overall_score) FROM quality_checks").fetchone()[0] or 0
        avg_nps = conn.execute("SELECT AVG(nps_score) FROM quality_checks WHERE nps_score > 0").fetchone()[0] or 0
        total = conn.execute("SELECT COUNT(*) FROM quality_checks").fetchone()[0]
        passed = conn.execute("SELECT COUNT(*) FROM quality_checks WHERE overall_score >= 80").fetchone()[0]
    return render_template("quality_dashboard.html", checks=checks, avg_score=avg_score, 
                          avg_nps=avg_nps, total=total, passed=passed)



# ─── 9. Tableau Comparatif Mensuel ───

@reports_bp.route("/monthly_comparison")
@login_required
@admin_required
def monthly_comparison_view():
    from datetime import date
    today = date.today()
    current_month = today.strftime("%Y-%m")
    if today.month == 1:
        prev_month = f"{today.year - 1}-12"
    else:
        prev_month = f"{today.year}-{today.month - 1:02d}"
    with get_db() as conn:
        def month_stats(m):
            revenue = conn.execute("SELECT COALESCE(SUM(amount),0) FROM invoices WHERE strftime('%%Y-%%m',created_at)=? AND status='paid'", (m,)).fetchone()[0]
            appts = conn.execute("SELECT COUNT(*) FROM appointments WHERE strftime('%%Y-%%m',date)=?", (m,)).fetchone()[0]
            completed = conn.execute("SELECT COUNT(*) FROM appointments WHERE strftime('%%Y-%%m',date)=? AND status='completed'", (m,)).fetchone()[0]
            new_customers = conn.execute("SELECT COUNT(*) FROM customers WHERE last_visit LIKE ?", (m + '%',)).fetchone()[0]
            expenses = conn.execute("SELECT COALESCE(SUM(amount),0) FROM expenses WHERE strftime('%%Y-%%m',date)=?", (m,)).fetchone()[0]
            avg_ticket = revenue / completed if completed else 0
            profit = revenue - expenses
            return {'month': m, 'revenue': revenue, 'appointments': appts, 'completed': completed,
                    'new_customers': new_customers, 'expenses': expenses, 'avg_ticket': avg_ticket, 'profit': profit}
        current = month_stats(current_month)
        previous = month_stats(prev_month)
        # Calculate deltas
        def delta(curr, prev):
            if prev == 0:
                return 100 if curr > 0 else 0
            return ((curr - prev) / prev) * 100
        deltas = {}
        for key in ['revenue', 'appointments', 'completed', 'new_customers', 'expenses', 'avg_ticket', 'profit']:
            deltas[key] = delta(current[key], previous[key])
        # Last 6 months for chart
        months_data = []
        for i in range(5, -1, -1):
            m = today.month - i
            y = today.year
            while m <= 0:
                m += 12; y -= 1
            ms = f"{y}-{m:02d}"
            months_data.append(month_stats(ms))
    return render_template("monthly_comparison.html", current=current, previous=previous,
                          deltas=deltas, months_data=months_data)



# ─── 6. Rentabilité par Type Véhicule ───

@reports_bp.route("/profitability_vehicle_type")
@login_required
@admin_required
def profitability_vehicle_type():
    month = request.args.get("month", "")
    from datetime import date
    if not month:
        month = date.today().strftime("%Y-%m")
    with get_db() as conn:
        data = conn.execute("""SELECT c.vehicle_type,
            COUNT(DISTINCT a.id) as appointments,
            COALESCE(SUM(CASE WHEN i.status='paid' THEN i.amount ELSE 0 END),0) as revenue,
            COALESCE(SUM(pu.total_cost),0) as product_cost,
            COUNT(DISTINCT c.id) as vehicles
            FROM appointments a
            JOIN cars c ON a.car_id=c.id
            LEFT JOIN invoices i ON i.appointment_id=a.id
            LEFT JOIN product_usage pu ON pu.appointment_id=a.id
            WHERE strftime('%%Y-%%m', a.date) = ?
            GROUP BY c.vehicle_type""", (month,)).fetchall()
        # Service breakdown by vehicle type
        service_data = conn.execute("""SELECT c.vehicle_type, a.service,
            COUNT(*) as cnt, COALESCE(SUM(CASE WHEN i.status='paid' THEN i.amount ELSE 0 END),0) as rev
            FROM appointments a JOIN cars c ON a.car_id=c.id LEFT JOIN invoices i ON i.appointment_id=a.id
            WHERE strftime('%%Y-%%m', a.date)=?
            GROUP BY c.vehicle_type, a.service ORDER BY rev DESC""", (month,)).fetchall()
    return render_template("profitability_vehicle_type.html", data=data, service_data=service_data, month=month)



@reports_bp.route("/efficiency_report")
@login_required
def efficiency_report():
    from datetime import date, timedelta
    month = request.args.get("month", date.today().strftime("%Y-%m"))
    with get_db() as conn:
        by_service = conn.execute("""SELECT service_name, COUNT(*) as count,
            AVG(estimated_minutes) as avg_est, AVG(actual_minutes) as avg_actual,
            AVG(efficiency_pct) as avg_eff
            FROM service_timer WHERE strftime('%%Y-%%m', created_at)=? AND actual_minutes > 0
            GROUP BY service_name ORDER BY avg_eff""", (month,)).fetchall()
        by_employee = conn.execute("""SELECT e.full_name as name, COUNT(*) as count,
            AVG(st.actual_minutes) as avg_time, AVG(st.efficiency_pct) as avg_eff
            FROM service_timer st LEFT JOIN users e ON st.employee_id=e.id
            WHERE strftime('%%Y-%%m', st.created_at)=? AND st.actual_minutes > 0
            GROUP BY st.employee_id ORDER BY avg_eff DESC""", (month,)).fetchall()
        bottlenecks = conn.execute("""SELECT service_name, employee_id, actual_minutes, estimated_minutes, efficiency_pct
            FROM service_timer WHERE efficiency_pct < 70 AND efficiency_pct > 0
            AND strftime('%%Y-%%m', created_at)=? ORDER BY efficiency_pct LIMIT 20""", (month,)).fetchall()
    return render_template("efficiency_report.html", by_service=by_service, by_employee=by_employee,
        bottlenecks=bottlenecks, month=month)



# ─── 8. Objectifs & Budget Mensuel ───

@reports_bp.route("/monthly_goals_view")
@login_required
def monthly_goals_view():
    from datetime import date
    month = request.args.get("month", date.today().strftime("%Y-%m"))
    with get_db() as conn:
        goals = conn.execute("SELECT * FROM monthly_goals WHERE month=? ORDER BY goal_type", (month,)).fetchall()
        actuals = {}
        actuals['revenue'] = conn.execute("""SELECT COALESCE(SUM(amount), 0) FROM invoices
            WHERE status='paid' AND strftime('%%Y-%%m', created_at)=?""", (month,)).fetchone()[0]
        actuals['appointments'] = conn.execute("""SELECT COUNT(*) FROM appointments
            WHERE strftime('%%Y-%%m', date)=?""", (month,)).fetchone()[0]
        actuals['new_customers'] = conn.execute("""SELECT COUNT(*) FROM customers
            WHERE strftime('%%Y-%%m', last_visit)=?""", (month,)).fetchone()[0]
        actuals['avg_ticket'] = conn.execute("""SELECT AVG(amount) FROM invoices
            WHERE status='paid' AND strftime('%%Y-%%m', created_at)=?""", (month,)).fetchone()[0] or 0
    return render_template("monthly_goals.html", goals=goals, actuals=actuals, month=month)



@reports_bp.route("/monthly_goal/add", methods=["POST"])
@login_required
def monthly_goal_add():
    with get_db() as conn:
        conn.execute("""INSERT INTO monthly_goals (month, goal_type, target_value, unit, notes)
            VALUES (?,?,?,?,?)""",
            (request.form["month"], request.form["goal_type"],
             request.form.get("target_value", 0, type=float),
             request.form.get("unit", ""), request.form.get("notes", "")))
        conn.commit()
    flash("Objectif ajouté", "success")
    return redirect(f"/monthly_goals_view?month={request.form['month']}")



@reports_bp.route('/smart_scheduling/suggest', methods=['POST'])
@login_required
def smart_scheduling_suggest():
    date = request.form['date']
    service_id = int(request.form.get('service_id', 0))
    duration = int(request.form.get('duration', 60))
    with get_db() as conn:
        # Find busy times
        busy = conn.execute("""
            SELECT time, estimated_duration FROM appointments WHERE date = ? AND status != 'cancelled'
        """, (date,)).fetchall()
        busy_times = set()
        for b in busy:
            if b['time']:
                hour = int(b['time'].split(':')[0]) if ':' in b['time'] else 8
                dur = b['estimated_duration'] or 60
                for h in range(hour, min(hour + (dur // 60) + 1, 19)):
                    busy_times.add(h)
        # Suggest free slots
        suggestions = []
        for hour in range(8, 18):
            slots_needed = max(1, duration // 60)
            if all(h not in busy_times for h in range(hour, min(hour + slots_needed, 19))):
                suggestions.append({'time': f"{hour:02d}:00", 'score': 100 - len(busy_times) * 5})
        # Sort by score
        suggestions.sort(key=lambda x: x['score'], reverse=True)
    return jsonify({'suggestions': suggestions[:5]})

# ── 8. Data Import Center ──
@reports_bp.route('/import_center')
@login_required
def import_center():
    with get_db() as conn:
        history = conn.execute("SELECT * FROM import_history ORDER BY created_at DESC LIMIT 20").fetchall()
    return render_template('import_center.html', history=history)



@reports_bp.route('/report_builder/create', methods=['POST'])
@login_required
def report_builder_create():
    sections = request.form.getlist('sections')
    with get_db() as conn:
        conn.execute("""INSERT INTO report_builder (name, report_type, sections, schedule)
            VALUES (?,?,?,?)""",
            (request.form['name'], request.form['report_type'],
             ','.join(sections), request.form.get('schedule', '')))
        conn.commit()
    flash("Rapport créé avec succès", "success")
    return redirect("/report_builder")



@reports_bp.route('/report_builder/generate/<int:report_id>')
@login_required
def report_builder_generate(report_id):
    with get_db() as conn:
        report = conn.execute("SELECT * FROM report_builder WHERE id=?", (report_id,)).fetchone()
        if not report:
            flash("Rapport non trouvé", "danger")
            return redirect("/report_builder")
        sections = (report['sections'] or '').split(',')
        data = {}
        month_start = datetime.now().strftime('%Y-%m-01')
        today = datetime.now().strftime('%Y-%m-%d')

        if 'revenue' in sections:
            data['revenue'] = conn.execute("""
                SELECT COALESCE(SUM(total), 0) as total, COUNT(*) as count,
                       AVG(total) as avg_ticket
                FROM invoices WHERE date >= ? AND status != 'cancelled'
            """, (month_start,)).fetchone()
        if 'appointments' in sections:
            data['appointments'] = conn.execute("""
                SELECT COUNT(*) as total,
                       SUM(CASE WHEN status='completed' THEN 1 ELSE 0 END) as completed,
                       SUM(CASE WHEN status='cancelled' THEN 1 ELSE 0 END) as cancelled
                FROM appointments WHERE date >= ?
            """, (month_start,)).fetchone()
        if 'customers' in sections:
            data['customers'] = conn.execute("""
                SELECT COUNT(*) as total,
                       SUM(CASE WHEN created_at >= ? THEN 1 ELSE 0 END) as new_this_month
                FROM customers
            """, (month_start,)).fetchone()
        if 'services' in sections:
            data['services'] = conn.execute("""
                SELECT service, COUNT(*) as count, SUM(i.total) as revenue
                FROM appointments a LEFT JOIN invoices i ON a.id = i.appointment_id
                WHERE a.date >= ? GROUP BY a.service ORDER BY count DESC LIMIT 10
            """, (month_start,)).fetchall()
        if 'employees' in sections:
            data['employees'] = conn.execute("""
                SELECT assigned_to, COUNT(*) as count
                FROM appointments WHERE date >= ? AND assigned_to != ''
                GROUP BY assigned_to ORDER BY count DESC
            """, (month_start,)).fetchall()
        if 'inventory' in sections:
            data['inventory'] = conn.execute("""
                SELECT name, quantity, min_quantity FROM inventory
                WHERE quantity <= min_quantity ORDER BY quantity ASC LIMIT 10
            """).fetchall()

        conn.execute("UPDATE report_builder SET last_generated=? WHERE id=?", (today, report_id))
        conn.commit()

    # Generate PDF
    settings = {}
    with get_db() as conn:
        for s in conn.execute("SELECT key, value FROM settings").fetchall():
            settings[s['key']] = s['value']

    html = render_template('report_builder_pdf.html', report=report, data=data,
                          sections=sections, settings=settings,
                          generated_at=datetime.now().strftime('%d/%m/%Y %H:%M'))
    from xhtml2pdf import pisa
    pdf_buffer = io.BytesIO()
    pisa.CreatePDF(io.BytesIO(html.encode('utf-8')), dest=pdf_buffer)
    pdf_buffer.seek(0)
    return send_file(pdf_buffer, mimetype='application/pdf',
                    download_name=f"rapport_{report['name']}_{today}.pdf", as_attachment=True)



@reports_bp.route('/report_builder/delete/<int:report_id>', methods=["POST"])
@login_required
def report_builder_delete(report_id):
    with get_db() as conn:
        conn.execute("DELETE FROM report_builder WHERE id=?", (report_id,))
        conn.commit()
    flash("Rapport supprimé", "success")
    return redirect("/report_builder")

# ── 10. Customer 360 View ──
@reports_bp.route('/customer_360/<int:customer_id>')
@login_required
def customer_360(customer_id):
    with get_db() as conn:
        customer = conn.execute("SELECT * FROM customers WHERE id=?", (customer_id,)).fetchone()
        if not customer:
            flash("Client non trouvé", "danger")
            return redirect("/customers")
        cars = conn.execute("SELECT * FROM cars WHERE customer_id=?", (customer_id,)).fetchall()
        appointments = conn.execute("""
            SELECT a.*, c.brand, c.model, c.plate FROM appointments a
            LEFT JOIN cars c ON a.car_id = c.id
            WHERE a.customer_id=? ORDER BY a.date DESC LIMIT 20
        """, (customer_id,)).fetchall()
        invoices = conn.execute("""
            SELECT * FROM invoices WHERE customer_id=? ORDER BY date DESC LIMIT 20
        """, (customer_id,)).fetchall()
        total_spent = conn.execute(
            "SELECT COALESCE(SUM(total), 0) FROM invoices WHERE customer_id=? AND status != 'cancelled'",
            (customer_id,)).fetchone()[0]
        total_visits = conn.execute(
            "SELECT COUNT(*) FROM appointments WHERE customer_id=?", (customer_id,)).fetchone()[0]
        # Wallet
        wallet = conn.execute(
            "SELECT * FROM wallet_transactions WHERE customer_id=? ORDER BY created_at DESC LIMIT 10",
            (customer_id,)).fetchall()
        # NPS
        nps = conn.execute(
            "SELECT * FROM nps_surveys WHERE customer_id=? ORDER BY created_at DESC LIMIT 5",
            (customer_id,)).fetchall()
        # Communications
        comms = conn.execute(
            "SELECT * FROM channel_inbox WHERE customer_id=? ORDER BY created_at DESC LIMIT 10",
            (customer_id,)).fetchall()
        # Loyalty
        loyalty = conn.execute(
            "SELECT * FROM loyalty WHERE customer_id=?", (customer_id,)).fetchone()
        # Treatments
        treatments = conn.execute("""
            SELECT t.* FROM treatments t
            LEFT JOIN cars c ON t.car_id = c.id
            WHERE c.customer_id=? ORDER BY t.applied_date DESC LIMIT 10
        """, (customer_id,)).fetchall()
        # Timeline
        timeline = conn.execute(
            "SELECT * FROM customer_timeline WHERE customer_id=? ORDER BY created_at DESC LIMIT 20",
            (customer_id,)).fetchall()
        # Reviews
        reviews = conn.execute(
            "SELECT * FROM client_reviews WHERE customer_id=? ORDER BY created_at DESC LIMIT 5",
            (customer_id,)).fetchall()
        # Referrals
        referrals = conn.execute(
            "SELECT * FROM referrals WHERE referrer_id=? OR referred_id=? ORDER BY created_at DESC",
            (customer_id, customer_id)).fetchall()
    return render_template('customer_360.html', customer=customer, cars=cars,
                          appointments=appointments, invoices=invoices,
                          total_spent=total_spent, total_visits=total_visits,
                          wallet=wallet, nps=nps, comms=comms, loyalty=loyalty,
                          treatments=treatments, timeline=timeline, reviews=reviews,
                          referrals=referrals)



# ─── 18.2 End of Day Report ──────────────────────────────────────────────────

@reports_bp.route("/end_of_day")
@login_required
def end_of_day_report():
    from datetime import date as d, timedelta
    day = request.args.get("date", str(d.today()))
    with get_db() as conn:
        appointments = conn.execute("""SELECT a.id, a.service, a.status, a.time, 
            c.name as customer_name, car.brand, car.model, car.plate
            FROM appointments a
            LEFT JOIN cars car ON a.car_id=car.id
            LEFT JOIN customers c ON car.customer_id=c.id
            WHERE a.date=? ORDER BY a.time""", (day,)).fetchall()
        
        revenue_paid = conn.execute("""SELECT COALESCE(SUM(i.amount),0) FROM invoices i
            JOIN appointments a ON i.appointment_id=a.id
            WHERE a.date=? AND i.status='paid'""", (day,)).fetchone()[0]
        revenue_unpaid = conn.execute("""SELECT COALESCE(SUM(i.amount),0) FROM invoices i
            JOIN appointments a ON i.appointment_id=a.id
            WHERE a.date=? AND i.status='unpaid'""", (day,)).fetchone()[0]
        expenses = conn.execute("SELECT COALESCE(SUM(amount),0) FROM expenses WHERE date=?", (day,)).fetchone()[0]
        
        stats = {
            'total_appointments': len(appointments),
            'completed': sum(1 for a in appointments if a['status'] == 'completed'),
            'in_progress': sum(1 for a in appointments if a['status'] == 'in_progress'),
            'pending': sum(1 for a in appointments if a['status'] == 'pending'),
            'cancelled': sum(1 for a in appointments if a['status'] == 'cancelled'),
            'revenue_paid': revenue_paid,
            'revenue_unpaid': revenue_unpaid,
            'expenses': expenses,
            'profit': revenue_paid - expenses,
        }
        
        by_service = conn.execute("""SELECT a.service, COUNT(*) as count,
            COALESCE(SUM(CASE WHEN i.status='paid' THEN i.amount ELSE 0 END),0) as revenue
            FROM appointments a LEFT JOIN invoices i ON a.id=i.appointment_id
            WHERE a.date=? GROUP BY a.service ORDER BY revenue DESC""", (day,)).fetchall()
        
        by_employee = conn.execute("""SELECT COALESCE(u.full_name, u.username, 'Non assigné') as name,
            COUNT(*) as count, SUM(CASE WHEN a.status='completed' THEN 1 ELSE 0 END) as completed
            FROM appointments a LEFT JOIN users u ON a.assigned_employee_id=u.id
            WHERE a.date=? GROUP BY a.assigned_employee_id""", (day,)).fetchall()
        
        payment_methods = conn.execute("""SELECT COALESCE(i.payment_method,'cash') as method,
            COUNT(*) as count, SUM(i.amount) as total
            FROM invoices i JOIN appointments a ON i.appointment_id=a.id
            WHERE a.date=? AND i.status='paid' GROUP BY i.payment_method""", (day,)).fetchall()
        
        shop = get_all_settings()
    return render_template("end_of_day.html", appointments=appointments, stats=stats,
                          by_service=by_service, by_employee=by_employee, 
                          payment_methods=payment_methods, day=day, shop=shop)





# ─── 18.6 Daily Technician Work Summary ─────────────────────────────────────

@reports_bp.route("/tech_summary")
@login_required
def tech_daily_summary():
    from datetime import date as d
    day = request.args.get("date", str(d.today()))
    with get_db() as conn:
        employees = conn.execute("SELECT id, full_name, username FROM users WHERE role IN ('employee','admin') ORDER BY full_name").fetchall()
        summaries = []
        for emp in employees:
            tasks = conn.execute("""SELECT a.service, a.status, a.time, 
                c.name as customer_name, car.brand, car.model, car.plate,
                COALESCE(i.amount,0) as revenue
                FROM appointments a
                LEFT JOIN cars car ON a.car_id=car.id
                LEFT JOIN customers c ON car.customer_id=c.id
                LEFT JOIN invoices i ON a.id=i.appointment_id
                WHERE a.date=? AND a.assigned_employee_id=?
                ORDER BY a.time""", (day, emp['id'])).fetchall()
            
            completed = sum(1 for t in tasks if t['status'] == 'completed')
            total_revenue = sum(t['revenue'] for t in tasks if t['status'] == 'completed')
            
            timer = conn.execute("""SELECT COUNT(*) as count, 
                COALESCE(AVG(efficiency_pct),0) as avg_eff,
                COALESCE(SUM(actual_minutes),0) as total_minutes
                FROM service_timer WHERE employee_id=? AND date(created_at)=?""",
                (emp['id'], day)).fetchone()
            
            summaries.append({
                'employee': emp,
                'tasks': tasks,
                'completed': completed,
                'total_tasks': len(tasks),
                'revenue': total_revenue,
                'avg_efficiency': round(timer['avg_eff'], 1) if timer else 0,
                'total_minutes': timer['total_minutes'] if timer else 0,
            })
    return render_template("tech_summary.html", summaries=summaries, day=day)


# ═══════════════════════════════════════════════════════════════════════════════
# EXPORT: Excel & PDF
# ═══════════════════════════════════════════════════════════════════════════════

@reports_bp.route('/export/customers_excel')
@login_required
def export_customers_excel():
    """Exporte la liste des clients en Excel (.xlsx)"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    import io

    with get_db() as conn:
        customers = conn.execute("""
            SELECT c.id, c.name, c.phone, c.email,
                   COALESCE(c.total_visits, 0) as visits,
                   COALESCE(c.last_visit, '') as last_visit,
                   COALESCE(rp.points, 0) as points,
                   COALESCE(rp.tier, 'Bronze') as tier,
                   COALESCE(SUM(i.amount), 0) as total_spent
            FROM customers c
            LEFT JOIN reward_points rp ON rp.customer_id = c.id
            LEFT JOIN cars ca ON ca.customer_id = c.id
            LEFT JOIN appointments a ON a.car_id = ca.id
            LEFT JOIN invoices i ON i.appointment_id = a.id AND i.status = 'paid'
            GROUP BY c.id ORDER BY total_spent DESC
        """).fetchall()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Clients"

    # Header style
    header_fill = PatternFill(start_color="1A1A2E", end_color="1A1A2E", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    border = Border(
        left=Side(style='thin', color='DDDDDD'),
        right=Side(style='thin', color='DDDDDD'),
        bottom=Side(style='thin', color='DDDDDD'),
    )

    headers = ["#", "Nom", "Téléphone", "Email", "Visites", "Dernière visite", "Points", "Tier", "CA Total (DT)"]
    col_widths = [5, 25, 15, 30, 8, 15, 8, 10, 15]

    # Title row
    ws.merge_cells("A1:I1")
    title_cell = ws["A1"]
    title_cell.value = "AMILCAR Auto Care — Liste des Clients"
    title_cell.font = Font(bold=True, size=14, color="1A1A2E")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    # Header row
    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[2].height = 22

    # Data rows
    tier_colors = {"Platinum": "E8C547", "Gold": "FFD700", "Silver": "C0C0C0", "Bronze": "CD7F32"}
    for row_idx, c in enumerate(customers, 3):
        row_fill = PatternFill(start_color="F8F9FA" if row_idx % 2 == 0 else "FFFFFF",
                               end_color="F8F9FA" if row_idx % 2 == 0 else "FFFFFF",
                               fill_type="solid")
        values = [c['id'], c['name'], c['phone'], c['email'] or '',
                  c['visits'], c['last_visit'], c['points'], c['tier'],
                  round(c['total_spent'], 2)]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.fill = row_fill
            cell.border = border
            cell.alignment = Alignment(vertical="center",
                                       horizontal="right" if col in (1, 5, 7, 9) else "left")
        # Color tier cell
        tier = c['tier']
        if tier in tier_colors:
            ws.cell(row=row_idx, column=8).fill = PatternFill(
                start_color=tier_colors[tier], end_color=tier_colors[tier], fill_type="solid")

    # Summary row
    last_row = len(customers) + 3
    ws.cell(row=last_row, column=1, value="TOTAL").font = Font(bold=True)
    ws.cell(row=last_row, column=5, value=f"=SUM(E3:E{last_row-1})").font = Font(bold=True)
    ws.cell(row=last_row, column=9, value=f"=SUM(I3:I{last_row-1})").font = Font(bold=True)

    # Freeze header
    ws.freeze_panes = "A3"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    log_activity('Export', 'Clients → Excel')
    return send_file(buf, as_attachment=True,
                     download_name=f"clients_amilcar_{__import__('datetime').date.today()}.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@reports_bp.route('/export/invoices_excel')
@login_required
def export_invoices_excel():
    """Exporte les factures en Excel"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    import io

    date_from = request.args.get('from', '')
    date_to = request.args.get('to', '')

    with get_db() as conn:
        q = """SELECT i.id, cu.name, a.date, a.service,
                      i.amount, i.paid_amount, i.status, i.payment_method,
                      ca.brand || ' ' || ca.model as car, ca.plate
               FROM invoices i
               JOIN appointments a ON i.appointment_id = a.id
               JOIN cars ca ON a.car_id = ca.id
               JOIN customers cu ON ca.customer_id = cu.id
               WHERE 1=1"""
        params = []
        if date_from:
            q += " AND a.date >= ?"
            params.append(date_from)
        if date_to:
            q += " AND a.date <= ?"
            params.append(date_to)
        q += " ORDER BY i.id DESC"
        invoices = conn.execute(q, params).fetchall()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Factures"

    header_fill = PatternFill(start_color="1A1A2E", end_color="1A1A2E", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    border = Border(
        left=Side(style='thin', color='DDDDDD'),
        right=Side(style='thin', color='DDDDDD'),
        bottom=Side(style='thin', color='DDDDDD'),
    )

    headers = ["N°", "Client", "Date", "Service", "Voiture", "Plaque", "Montant DT", "Payé DT", "Reste DT", "Statut", "Paiement"]
    col_widths = [6, 22, 12, 28, 20, 12, 12, 12, 12, 10, 12]

    ws.merge_cells("A1:K1")
    ws["A1"].value = f"AMILCAR — Factures{'  ' + date_from + ' → ' + date_to if date_from else ''}"
    ws["A1"].font = Font(bold=True, size=13, color="1A1A2E")
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 28

    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = border
        ws.column_dimensions[get_column_letter(col)].width = w

    status_colors = {"paid": "C8E6C9", "unpaid": "FFCDD2", "partial": "FFF9C4", "cancelled": "EEEEEE"}

    total_amount = total_paid = 0
    for row_idx, inv in enumerate(invoices, 3):
        amount = inv['amount'] or 0
        paid = inv['paid_amount'] or 0
        reste = round(amount - paid, 2)
        total_amount += amount
        total_paid += paid
        row_fill = PatternFill(start_color="F8F9FA" if row_idx % 2 == 0 else "FFFFFF",
                               end_color="F8F9FA" if row_idx % 2 == 0 else "FFFFFF",
                               fill_type="solid")
        values = [inv['id'], inv['name'], inv['date'], inv['service'],
                  inv['car'], inv['plate'], round(amount, 2), round(paid, 2), reste,
                  inv['status'], inv['payment_method'] or '']
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.fill = row_fill
            cell.border = border
            cell.alignment = Alignment(vertical="center",
                                       horizontal="right" if col in (7, 8, 9) else "left")
        # Color status
        st = inv['status']
        if st in status_colors:
            ws.cell(row=row_idx, column=10).fill = PatternFill(
                start_color=status_colors[st], end_color=status_colors[st], fill_type="solid")

    # Totals
    last = len(invoices) + 3
    for col, val in [(1, "TOTAL"), (7, round(total_amount, 2)),
                     (8, round(total_paid, 2)), (9, round(total_amount - total_paid, 2))]:
        cell = ws.cell(row=last, column=col, value=val)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")

    ws.freeze_panes = "A3"
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    log_activity('Export', 'Factures → Excel')
    return send_file(buf, as_attachment=True,
                     download_name=f"factures_amilcar_{__import__('datetime').date.today()}.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@reports_bp.route('/export/appointments_excel')
@login_required
def export_appointments_excel():
    """Exporte les rendez-vous en Excel"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    import io

    date_from = request.args.get('from', '')
    date_to = request.args.get('to', '')

    with get_db() as conn:
        q = """SELECT a.id, cu.name, cu.phone, ca.brand || ' ' || ca.model as car,
                      ca.plate, a.date, a.time, a.service, a.status
               FROM appointments a
               JOIN cars ca ON a.car_id = ca.id
               JOIN customers cu ON ca.customer_id = cu.id
               WHERE 1=1"""
        params = []
        if date_from:
            q += " AND a.date >= ?"
            params.append(date_from)
        if date_to:
            q += " AND a.date <= ?"
            params.append(date_to)
        q += " ORDER BY a.date DESC, a.time"
        appts = conn.execute(q, params).fetchall()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Rendez-vous"

    header_fill = PatternFill(start_color="0F3460", end_color="0F3460", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    border = Border(
        left=Side(style='thin', color='DDDDDD'),
        right=Side(style='thin', color='DDDDDD'),
        bottom=Side(style='thin', color='DDDDDD'),
    )

    headers = ["#", "Client", "Téléphone", "Véhicule", "Plaque", "Date", "Heure", "Service", "Statut"]
    col_widths = [5, 22, 14, 20, 12, 12, 8, 28, 12]

    ws.merge_cells("A1:I1")
    ws["A1"].value = "AMILCAR Auto Care — Rendez-vous"
    ws["A1"].font = Font(bold=True, size=13, color="0F3460")
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 28

    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = border
        ws.column_dimensions[get_column_letter(col)].width = w

    status_colors = {
        "completed": "C8E6C9", "pending": "FFF9C4",
        "cancelled": "FFCDD2", "in_progress": "BBDEFB"
    }
    for row_idx, a in enumerate(appts, 3):
        row_fill = PatternFill(start_color="F8F9FA" if row_idx % 2 == 0 else "FFFFFF",
                               end_color="F8F9FA" if row_idx % 2 == 0 else "FFFFFF",
                               fill_type="solid")
        values = [a['id'], a['name'], a['phone'], a['car'], a['plate'],
                  a['date'], a['time'] or '', a['service'], a['status']]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.fill = row_fill
            cell.border = border
            cell.alignment = Alignment(vertical="center")
        st = a['status']
        if st in status_colors:
            ws.cell(row=row_idx, column=9).fill = PatternFill(
                start_color=status_colors[st], end_color=status_colors[st], fill_type="solid")

    ws.freeze_panes = "A3"
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    log_activity('Export', 'RDV → Excel')
    return send_file(buf, as_attachment=True,
                     download_name=f"rdv_amilcar_{__import__('datetime').date.today()}.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@reports_bp.route('/export/full_report_excel')
@reports_bp.route('/export/monthly')
@login_required
def export_monthly_report_excel():
    """Rapport mensuel complet en Excel avec plusieurs feuilles"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.chart import BarChart, Reference
    from openpyxl.utils import get_column_letter
    import io
    from datetime import date

    month = request.args.get('month', date.today().strftime('%Y-%m'))

    with get_db() as conn:
        invoices = conn.execute("""
            SELECT i.id, cu.name, a.date, a.service, i.amount, i.status, i.payment_method
            FROM invoices i JOIN appointments a ON i.appointment_id=a.id
            JOIN cars ca ON a.car_id=ca.id JOIN customers cu ON ca.customer_id=cu.id
            WHERE strftime('%Y-%m', a.date) = ? ORDER BY a.date
        """, (month,)).fetchall()

        expenses = conn.execute("""
            SELECT id, description, amount, category, date
            FROM expenses WHERE strftime('%Y-%m', date) = ? ORDER BY date
        """, (month,)).fetchall()

        top_services = conn.execute("""
            SELECT a.service, COUNT(*) as cnt, COALESCE(SUM(i.amount),0) as revenue
            FROM appointments a LEFT JOIN invoices i ON a.id=i.appointment_id
            WHERE strftime('%Y-%m', a.date) = ? AND a.status='completed'
            GROUP BY a.service ORDER BY revenue DESC LIMIT 10
        """, (month,)).fetchall()

        new_customers = conn.execute(
            "SELECT COUNT(*) FROM customers WHERE strftime('%Y-%m', last_visit) = ?",
            (month,)).fetchone()[0]

    wb = openpyxl.Workbook()
    header_fill = PatternFill(start_color="1A1A2E", end_color="1A1A2E", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    border = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC'),
    )

    def style_header(ws, row, cols, headers, widths):
        for col, (h, w) in enumerate(zip(headers, widths), 1):
            cell = ws.cell(row=row, column=col, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
            cell.border = border
            ws.column_dimensions[get_column_letter(col)].width = w

    # ── Feuille 1: Résumé ──────────────────────────────────────────────────
    ws_sum = wb.active
    ws_sum.title = "Résumé"
    total_revenue = sum(r['amount'] for r in invoices if r['status'] == 'paid')
    total_invoices = len(invoices)
    total_expenses = sum(e['amount'] for e in expenses)
    profit = total_revenue - total_expenses

    summary_data = [
        ("Mois", month),
        ("Chiffre d'affaires (DT)", round(total_revenue, 2)),
        ("Nombre de factures", total_invoices),
        ("Dépenses totales (DT)", round(total_expenses, 2)),
        ("Bénéfice net (DT)", round(profit, 2)),
        ("Nouveaux clients", new_customers),
        ("Taux de marge (%)", round((profit / total_revenue * 100) if total_revenue else 0, 1)),
    ]
    ws_sum.merge_cells("A1:B1")
    ws_sum["A1"].value = f"AMILCAR — Rapport Mensuel {month}"
    ws_sum["A1"].font = Font(bold=True, size=14, color="1A1A2E")
    ws_sum["A1"].alignment = Alignment(horizontal="center")
    ws_sum.row_dimensions[1].height = 30
    ws_sum.column_dimensions["A"].width = 30
    ws_sum.column_dimensions["B"].width = 20

    for r, (label, val) in enumerate(summary_data, 2):
        ws_sum.cell(row=r, column=1, value=label).font = Font(bold=True)
        cell = ws_sum.cell(row=r, column=2, value=val)
        if "Bénéfice" in label:
            cell.font = Font(bold=True, color="2E7D32" if profit >= 0 else "C62828")
        ws_sum.row_dimensions[r].height = 20

    # ── Feuille 2: Factures ────────────────────────────────────────────────
    ws_inv = wb.create_sheet("Factures")
    ws_inv.merge_cells("A1:G1")
    ws_inv["A1"].value = f"Factures — {month}"
    ws_inv["A1"].font = Font(bold=True, size=12, color="1A1A2E")
    ws_inv["A1"].alignment = Alignment(horizontal="center")
    style_header(ws_inv, 2,
                 ["#", "Client", "Date", "Service", "Montant DT", "Statut", "Paiement"],
                 ["#", "Client", "Date", "Service", "Montant DT", "Statut", "Paiement"],
                 [5, 22, 12, 28, 12, 10, 12])
    for ri, inv in enumerate(invoices, 3):
        row_fill = PatternFill(start_color="F8F9FA" if ri % 2 == 0 else "FFFFFF",
                               end_color="F8F9FA" if ri % 2 == 0 else "FFFFFF", fill_type="solid")
        for ci, val in enumerate([inv['id'], inv['name'], inv['date'], inv['service'],
                                   round(inv['amount'], 2), inv['status'], inv['payment_method'] or ''], 1):
            c = ws_inv.cell(row=ri, column=ci, value=val)
            c.fill = row_fill
            c.border = border

    # ── Feuille 3: Top Services ────────────────────────────────────────────
    ws_svc = wb.create_sheet("Top Services")
    ws_svc.merge_cells("A1:C1")
    ws_svc["A1"].value = f"Top Services — {month}"
    ws_svc["A1"].font = Font(bold=True, size=12, color="1A1A2E")
    ws_svc["A1"].alignment = Alignment(horizontal="center")
    style_header(ws_svc, 2,
                 ["Service", "Nombre", "CA (DT)"],
                 ["Service", "Nombre", "CA (DT)"],
                 [30, 10, 14])
    for ri, svc in enumerate(top_services, 3):
        for ci, val in enumerate([svc['service'], svc['cnt'], round(svc['revenue'], 2)], 1):
            c = ws_svc.cell(row=ri, column=ci, value=val)
            c.border = border

    # ── Feuille 4: Dépenses ────────────────────────────────────────────────
    ws_exp = wb.create_sheet("Dépenses")
    ws_exp.merge_cells("A1:E1")
    ws_exp["A1"].value = f"Dépenses — {month}"
    ws_exp["A1"].font = Font(bold=True, size=12, color="1A1A2E")
    ws_exp["A1"].alignment = Alignment(horizontal="center")
    style_header(ws_exp, 2,
                 ["#", "Description", "Montant DT", "Catégorie", "Date"],
                 ["#", "Description", "Montant DT", "Catégorie", "Date"],
                 [5, 30, 14, 16, 12])
    for ri, exp in enumerate(expenses, 3):
        for ci, val in enumerate([exp['id'], exp['description'],
                                   round(exp['amount'], 2), exp['category'], exp['date']], 1):
            c = ws_exp.cell(row=ri, column=ci, value=val)
            c.border = border

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    log_activity('Export', f'Rapport mensuel {month} → Excel')
    return send_file(buf, as_attachment=True,
                     download_name=f"rapport_{month}_amilcar.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@reports_bp.route('/export/inventory_excel')
@login_required
def export_inventory_excel():
    """Exporte l'inventaire en Excel"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    import io

    with get_db() as conn:
        items = conn.execute(
            "SELECT id, name, category, quantity, min_quantity, unit_price FROM inventory ORDER BY category, name"
        ).fetchall()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Inventaire"

    header_fill = PatternFill(start_color="0F3460", end_color="0F3460", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    border = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC'),
    )

    ws.merge_cells("A1:F1")
    ws["A1"].value = f"AMILCAR — Inventaire au {__import__('datetime').date.today()}"
    ws["A1"].font = Font(bold=True, size=13, color="0F3460")
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 28

    headers = ["#", "Article", "Catégorie", "Quantité", "Qté Min", "Prix Unitaire DT"]
    col_widths = [5, 30, 16, 10, 10, 18]
    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = border
        ws.column_dimensions[get_column_letter(col)].width = w

    low_stock_fill = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")
    ok_stock_fill  = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")

    for row_idx, item in enumerate(items, 3):
        is_low = (item['quantity'] or 0) <= (item['min_quantity'] or 0)
        row_fill = low_stock_fill if is_low else (
            PatternFill(start_color="F8F9FA" if row_idx % 2 == 0 else "FFFFFF",
                        end_color="F8F9FA" if row_idx % 2 == 0 else "FFFFFF",
                        fill_type="solid"))
        values = [item['id'], item['name'], item['category'],
                  item['quantity'], item['min_quantity'],
                  round(item['unit_price'] or 0, 2)]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.fill = row_fill
            cell.border = border
            cell.alignment = Alignment(vertical="center",
                                       horizontal="right" if col in (4, 5, 6) else "left")

    # Legend
    last = len(items) + 4
    ws.cell(row=last, column=1, value="🔴 Rouge = stock bas").font = Font(italic=True, color="C62828")
    ws.cell(row=last+1, column=1, value="🟢 Vert = stock OK").font = Font(italic=True, color="2E7D32")

    ws.freeze_panes = "A3"
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    log_activity('Export', 'Inventaire → Excel')
    return send_file(buf, as_attachment=True,
                     download_name=f"inventaire_amilcar_{__import__('datetime').date.today()}.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


# ── PDF Exports ──────────────────────────────────────────────────────────────

def _render_pdf(html_content):
    """Helper: convert HTML to PDF bytes"""
    from xhtml2pdf import pisa
    import io
    buf = io.BytesIO()
    pisa.CreatePDF(io.StringIO(html_content), dest=buf)
    buf.seek(0)
    return buf


@reports_bp.route('/export/monthly_pdf')
@login_required
def export_monthly_pdf():
    """Rapport mensuel en PDF"""
    from datetime import date as dt_date
    month = request.args.get('month', dt_date.today().strftime('%Y-%m'))

    with get_db() as conn:
        invoices = conn.execute("""
            SELECT i.id, cu.name, a.date, a.service, i.amount, i.status, i.payment_method
            FROM invoices i JOIN appointments a ON i.appointment_id=a.id
            JOIN cars ca ON a.car_id=ca.id JOIN customers cu ON ca.customer_id=cu.id
            WHERE strftime('%Y-%m', a.date) = ? ORDER BY a.date
        """, (month,)).fetchall()

        expenses = conn.execute(
            "SELECT description, amount, category FROM expenses WHERE strftime('%Y-%m', date) = ? ORDER BY date",
            (month,)).fetchall()

        top_services = conn.execute("""
            SELECT a.service, COUNT(*) as cnt, COALESCE(SUM(i.amount),0) as revenue
            FROM appointments a LEFT JOIN invoices i ON a.id=i.appointment_id
            WHERE strftime('%Y-%m', a.date) = ? AND a.status='completed'
            GROUP BY a.service ORDER BY revenue DESC LIMIT 8
        """, (month,)).fetchall()

        settings = get_all_settings()

    total_revenue = sum(r['amount'] for r in invoices if r['status'] == 'paid')
    total_expenses_sum = sum(e['amount'] for e in expenses)
    profit = total_revenue - total_expenses_sum
    shop_name = settings.get('shop_name', 'AMILCAR Auto Care')

    profit_color = '#2e7d32' if profit >= 0 else '#c62828'
    rows_invoices = ''.join(
        f"<tr><td>{r['id']}</td><td>{r['name']}</td><td>{r['date']}</td>"
        f"<td>{r['service']}</td><td>{round(r['amount'],2)} DT</td>"
        f"<td class='{r['status']}'>{r['status']}</td></tr>"
        for r in invoices[:50]
    )
    rows_services = ''.join(
        f"<tr><td>{s['service']}</td><td>{s['cnt']}</td><td>{round(s['revenue'],2)}</td></tr>"
        for s in top_services
    )
    rows_expenses = ''.join(
        f"<tr><td>{e['description']}</td><td>{round(e['amount'],2)} DT</td><td>{e['category']}</td></tr>"
        for e in expenses
    )

    html = f"""<!DOCTYPE html>
<html><head><meta charset="utf-8">
<style>
  body {{ font-family: Arial, sans-serif; font-size: 11px; color: #1a1a2e; margin: 20px; }}
  h1 {{ color: #1a1a2e; font-size: 20px; text-align: center; border-bottom: 2px solid #e8c547; padding-bottom: 8px; }}
  h2 {{ color: #0f3460; font-size: 14px; margin-top: 20px; border-left: 4px solid #e8c547; padding-left: 8px; }}
  .kpi-grid {{ display: table; width: 100%; margin: 16px 0; }}
  .kpi {{ display: table-cell; text-align: center; padding: 10px; background: #f8f9fa; border: 1px solid #ddd; }}
  .kpi-val {{ font-size: 20px; font-weight: bold; color: #0f3460; }}
  .kpi-label {{ font-size: 10px; color: #666; }}
  table {{ width: 100%; border-collapse: collapse; margin-top: 8px; }}
  th {{ background: #1a1a2e; color: white; padding: 6px 8px; text-align: left; font-size: 10px; }}
  td {{ padding: 5px 8px; border-bottom: 1px solid #eee; font-size: 10px; }}
  tr:nth-child(even) td {{ background: #f8f9fa; }}
  .paid {{ color: #2e7d32; font-weight: bold; }}
  .unpaid {{ color: #c62828; font-weight: bold; }}
  .partial {{ color: #e65100; font-weight: bold; }}
  .footer {{ text-align: center; color: #999; font-size: 9px; margin-top: 30px; border-top: 1px solid #eee; padding-top: 8px; }}
</style>
</head><body>
<h1>AMILCAR — Rapport Mensuel {month}</h1>
<div class="kpi-grid">
  <div class="kpi"><div class="kpi-val">{round(total_revenue, 0):.0f} DT</div><div class="kpi-label">Chiffre d&apos;affaires</div></div>
  <div class="kpi"><div class="kpi-val">{len(invoices)}</div><div class="kpi-label">Factures</div></div>
  <div class="kpi"><div class="kpi-val">{round(total_expenses_sum, 0):.0f} DT</div><div class="kpi-label">Depenses</div></div>
  <div class="kpi"><div class="kpi-val" style="color:{profit_color}">{round(profit, 0):.0f} DT</div><div class="kpi-label">Benefice net</div></div>
</div>
<h2>Top Services</h2>
<table><tr><th>Service</th><th>Nombre</th><th>Revenu (DT)</th></tr>{rows_services}</table>
<h2>Factures ({len(invoices)})</h2>
<table><tr><th>#</th><th>Client</th><th>Date</th><th>Service</th><th>Montant</th><th>Statut</th></tr>{rows_invoices}</table>
<h2>Depenses ({len(expenses)})</h2>
<table><tr><th>Description</th><th>Montant</th><th>Categorie</th></tr>{rows_expenses}</table>
<div class="footer">Genere le {dt_date.today()} — {shop_name}</div>
</body></html>"""

    buf = _render_pdf(html)
    log_activity('Export', f'Rapport {month} → PDF')
    return send_file(buf, as_attachment=True,
                     download_name=f"rapport_{month}_amilcar.pdf",
                     mimetype="application/pdf")


@reports_bp.route('/export/professional_pdf')
@login_required
def export_professional_pdf():
    """Rapport general professionnel PDF"""
    from datetime import date as dt_date

    today = dt_date.today()
    month_start = today.strftime('%Y-%m-01')

    with get_db() as conn:
        settings = get_all_settings()
        total_customers = conn.execute("SELECT COUNT(*) FROM customers").fetchone()[0]
        total_cars = conn.execute("SELECT COUNT(*) FROM cars").fetchone()[0]
        revenue_month = conn.execute(
            "SELECT COALESCE(SUM(amount),0) FROM invoices WHERE status='paid' AND created_at >= ?",
            (month_start,)).fetchone()[0]
        revenue_total = conn.execute(
            "SELECT COALESCE(SUM(amount),0) FROM invoices WHERE status='paid'").fetchone()[0]
        appts_total = conn.execute("SELECT COUNT(*) FROM appointments").fetchone()[0]
        appts_completed = conn.execute(
            "SELECT COUNT(*) FROM appointments WHERE status='completed'").fetchone()[0]
        top_services = conn.execute("""
            SELECT a.service, COUNT(*) as cnt, COALESCE(SUM(i.amount),0) as rev
            FROM appointments a LEFT JOIN invoices i ON a.id=i.appointment_id
            WHERE a.status='completed' GROUP BY a.service ORDER BY rev DESC LIMIT 8
        """).fetchall()
        recent_invoices = conn.execute("""
            SELECT i.id, cu.name, a.date, a.service, i.amount, i.status
            FROM invoices i JOIN appointments a ON i.appointment_id=a.id
            JOIN cars ca ON a.car_id=ca.id JOIN customers cu ON ca.customer_id=cu.id
            ORDER BY i.id DESC LIMIT 15
        """).fetchall()
        low_stock = conn.execute(
            "SELECT name, quantity, min_quantity FROM inventory WHERE quantity <= min_quantity AND min_quantity > 0"
        ).fetchall()

    shop_name = settings.get('shop_name', 'AMILCAR Auto Care')
    completion_rate = round(appts_completed / appts_total * 100, 1) if appts_total else 0

    rows_services = ''.join(
        f"<tr><td>{s['service']}</td><td>{s['cnt']}</td><td>{round(s['rev'],2)}</td></tr>"
        for s in top_services
    )
    rows_invoices = ''.join(
        f"<tr><td>{r['id']}</td><td>{r['name']}</td><td>{r['date']}</td>"
        f"<td>{r['service']}</td><td>{round(r['amount'],2)}</td>"
        f"<td class='badge-{r['status']}'>{r['status']}</td></tr>"
        for r in recent_invoices
    )
    stock_alerts = ''.join(
        f"<p>- {i['name']} : {i['quantity']} unites (min: {i['min_quantity']})</p>"
        for i in low_stock
    )
    stock_section = (
        f'<h2>Alertes Stock</h2><div class="alert-box">{stock_alerts}</div>'
        if low_stock else ''
    )

    html = f"""<!DOCTYPE html>
<html><head><meta charset="utf-8">
<style>
  body {{ font-family: Arial, sans-serif; font-size: 11px; color: #1a1a2e; margin: 24px; }}
  .header {{ background: #1a1a2e; color: white; padding: 16px; text-align: center; }}
  .header h1 {{ color: #e8c547; margin: 0; font-size: 20px; }}
  .header p {{ color: #aaa; margin: 4px 0 0; }}
  h2 {{ color: #0f3460; font-size: 13px; margin-top: 22px; border-bottom: 2px solid #e8c547; padding-bottom: 4px; }}
  .kpi-row {{ display: table; width: 100%; margin: 12px 0; }}
  .kpi {{ display: table-cell; padding: 10px; background: #f0f4ff; border: 1px solid #c5cef0; text-align: center; }}
  .kpi-val {{ font-size: 20px; font-weight: bold; color: #0f3460; }}
  table {{ width: 100%; border-collapse: collapse; margin-top: 8px; font-size: 10px; }}
  th {{ background: #0f3460; color: white; padding: 6px 8px; text-align: left; }}
  td {{ padding: 5px 8px; border-bottom: 1px solid #eee; }}
  tr:nth-child(even) td {{ background: #f8f9fa; }}
  .badge-paid {{ color: #2e7d32; font-weight: bold; }}
  .badge-unpaid {{ color: #c62828; }}
  .badge-partial {{ color: #e65100; }}
  .alert-box {{ background: #fff3cd; border: 1px solid #ffc107; padding: 10px; margin: 10px 0; }}
  .footer {{ text-align: center; color: #aaa; font-size: 9px; margin-top: 30px; border-top: 1px solid #ddd; padding-top: 8px; }}
</style>
</head><body>
<div class="header">
  <h1>{shop_name}</h1>
  <p>Rapport Professionnel — {today.strftime('%d/%m/%Y')}</p>
</div>
<h2>Indicateurs Cles</h2>
<div class="kpi-row">
  <div class="kpi"><div class="kpi-val">{total_customers}</div><div>Clients</div></div>
  <div class="kpi"><div class="kpi-val">{total_cars}</div><div>Vehicules</div></div>
  <div class="kpi"><div class="kpi-val">{appts_total}</div><div>RDV total</div></div>
  <div class="kpi"><div class="kpi-val">{completion_rate}%</div><div>Taux completion</div></div>
</div>
<div class="kpi-row">
  <div class="kpi"><div class="kpi-val">{round(revenue_month):,} DT</div><div>CA ce mois</div></div>
  <div class="kpi"><div class="kpi-val">{round(revenue_total):,} DT</div><div>CA total</div></div>
</div>
<h2>Top Services</h2>
<table><tr><th>Service</th><th>Prestations</th><th>Revenu (DT)</th></tr>{rows_services}</table>
<h2>Dernieres Factures</h2>
<table><tr><th>#</th><th>Client</th><th>Date</th><th>Service</th><th>Montant DT</th><th>Statut</th></tr>{rows_invoices}</table>
{stock_section}
<div class="footer">{shop_name} — Rapport confidentiel — {today}</div>
</body></html>"""

    buf = _render_pdf(html)
    log_activity('Export', 'Rapport professionnel → PDF')
    return send_file(buf, as_attachment=True,
                     download_name=f"rapport_professionnel_{today}.pdf",
                     mimetype="application/pdf")


@reports_bp.route('/export/daily')
@login_required
def export_daily_pdf():
    """Rapport journalier PDF"""
    from datetime import date as dt_date

    day = request.args.get('date', dt_date.today().isoformat())

    with get_db() as conn:
        settings = get_all_settings()
        appts = conn.execute("""
            SELECT a.id, cu.name, cu.phone, ca.brand || ' ' || ca.model as car,
                   ca.plate, a.time, a.service, a.status,
                   COALESCE(i.amount, 0) as amount
            FROM appointments a
            JOIN cars ca ON a.car_id = ca.id
            JOIN customers cu ON ca.customer_id = cu.id
            LEFT JOIN invoices i ON i.appointment_id = a.id
            WHERE a.date = ? ORDER BY a.time
        """, (day,)).fetchall()

        revenue = conn.execute(
            "SELECT COALESCE(SUM(i.amount),0) FROM invoices i "
            "JOIN appointments a ON i.appointment_id=a.id WHERE a.date=? AND i.status='paid'",
            (day,)).fetchone()[0]

    shop_name = settings.get('shop_name', 'AMILCAR Auto Care')
    completed = sum(1 for a in appts if a['status'] == 'completed')

    rows = ''.join(
        f"<tr><td>{a['time'] or '-'}</td><td>{a['name']}</td><td>{a['car']}</td>"
        f"<td>{a['plate']}</td><td>{a['service']}</td>"
        f"<td class='{a['status']}'>{a['status']}</td>"
        f"<td>{round(a['amount'],2) if a['amount'] else '-'} DT</td></tr>"
        for a in appts
    )

    html = f"""<!DOCTYPE html>
<html><head><meta charset="utf-8">
<style>
  body {{ font-family: Arial, sans-serif; font-size: 11px; color: #1a1a2e; margin: 24px; }}
  .header {{ background: #1a1a2e; color: white; padding: 14px; text-align: center; }}
  .header h1 {{ color: #e8c547; margin: 0; font-size: 18px; }}
  .stats {{ display: table; width: 100%; margin: 14px 0; }}
  .stat {{ display: table-cell; text-align: center; padding: 10px; background: #f0f4ff; border: 1px solid #ddd; }}
  .stat-val {{ font-size: 20px; font-weight: bold; color: #0f3460; }}
  table {{ width: 100%; border-collapse: collapse; margin-top: 12px; }}
  th {{ background: #0f3460; color: white; padding: 7px; font-size: 10px; }}
  td {{ padding: 6px 7px; border-bottom: 1px solid #eee; font-size: 10px; }}
  tr:nth-child(even) td {{ background: #f8f9fa; }}
  .completed {{ color: #2e7d32; font-weight: bold; }}
  .pending {{ color: #e65100; }}
  .cancelled {{ color: #c62828; }}
</style>
</head><body>
<div class="header">
  <h1>{shop_name} — Rapport Journalier</h1>
  <p style="color:#aaa;margin:4px 0 0">{day}</p>
</div>
<div class="stats">
  <div class="stat"><div class="stat-val">{len(appts)}</div><div>RDV total</div></div>
  <div class="stat"><div class="stat-val">{completed}</div><div>Completes</div></div>
  <div class="stat"><div class="stat-val">{round(revenue,0):.0f} DT</div><div>Revenu</div></div>
</div>
<table>
  <tr><th>Heure</th><th>Client</th><th>Vehicule</th><th>Plaque</th><th>Service</th><th>Statut</th><th>Montant</th></tr>
  {rows}
</table>
</body></html>"""

    buf = _render_pdf(html)
    log_activity('Export', f'Rapport journalier {day} → PDF')
    return send_file(buf, as_attachment=True,
                     download_name=f"rapport_journalier_{day}.pdf",
                     mimetype="application/pdf")


@reports_bp.route('/export/expenses')
@login_required
def export_expenses_excel():
    """Exporte les depenses en Excel"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    import io
    from datetime import date as dt_date

    month = request.args.get('month', dt_date.today().strftime('%Y-%m'))

    with get_db() as conn:
        expenses = conn.execute(
            "SELECT id, description, amount, category, date FROM expenses "
            "WHERE strftime('%Y-%m', date) = ? ORDER BY date",
            (month,)).fetchall()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Depenses"

    header_fill = PatternFill(start_color="C62828", end_color="C62828", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    border = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC'),
    )

    ws.merge_cells("A1:E1")
    ws["A1"].value = f"AMILCAR — Depenses {month}"
    ws["A1"].font = Font(bold=True, size=13, color="C62828")
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 28

    headers = ["#", "Description", "Montant DT", "Categorie", "Date"]
    col_widths = [5, 35, 14, 18, 12]
    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = border
        ws.column_dimensions[get_column_letter(col)].width = w

    total = 0
    for row_idx, exp in enumerate(expenses, 3):
        total += exp['amount'] or 0
        row_fill = PatternFill(start_color="F8F9FA" if row_idx % 2 == 0 else "FFFFFF",
                               end_color="F8F9FA" if row_idx % 2 == 0 else "FFFFFF",
                               fill_type="solid")
        for col, val in enumerate([exp['id'], exp['description'],
                                    round(exp['amount'], 2), exp['category'], exp['date']], 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.fill = row_fill
            cell.border = border
            cell.alignment = Alignment(horizontal="right" if col == 3 else "left")

    last_row = len(expenses) + 3
    ws.cell(row=last_row, column=2, value="TOTAL").font = Font(bold=True)
    total_cell = ws.cell(row=last_row, column=3, value=round(total, 2))
    total_cell.font = Font(bold=True, color="C62828")
    total_cell.fill = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")

    ws.freeze_panes = "A3"
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    log_activity('Export', f'Depenses {month} → Excel')
    return send_file(buf, as_attachment=True,
                     download_name=f"depenses_{month}_amilcar.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
