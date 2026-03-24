"""
AMILCAR — Data Export (Excel, PDF, CSV)
Blueprint: exports_bp
"""
from flask import Blueprint, render_template, request, redirect, flash, make_response, jsonify, session, send_file
from helpers import login_required, admin_required, get_db, get_services, get_setting, get_all_settings
from helpers import log_activity, PER_PAGE
from database.db import get_db
from datetime import datetime, date, timedelta
import os, io
import time as time_module

exports_bp = Blueprint("exports_bp", __name__)

# ─── 2. Export Data (Excel/CSV/PDF) ───
@exports_bp.route("/export_data")
@login_required
def export_data():
    return render_template("export_data.html")



@exports_bp.route("/export_data/<data_type>/<fmt>")
@login_required
def export_data_download(data_type, fmt):
    import csv
    allowed_types = ['customers', 'invoices', 'appointments', 'cars', 'expenses']
    allowed_fmts = ['csv', 'pdf']
    if data_type not in allowed_types or fmt not in allowed_fmts:
        flash("Type ou format non supporté", "error")
        return redirect("/export_data")

    with get_db() as conn:
        if data_type == 'customers':
            rows = conn.execute("SELECT id, name, phone, email, notes FROM customers ORDER BY name").fetchall()
            headers = ['ID', 'Nom', 'Téléphone', 'Email', 'Notes']
        elif data_type == 'invoices':
            rows = conn.execute("""SELECT i.id, c.name, ca.plate, a.service, i.amount, i.status, i.payment_method
                FROM invoices i JOIN appointments a ON i.appointment_id=a.id
                JOIN cars ca ON a.car_id=ca.id JOIN customers c ON ca.customer_id=c.id
                ORDER BY i.id DESC""").fetchall()
            headers = ['ID', 'Client', 'Plaque', 'Service', 'Montant', 'Statut', 'Paiement']
        elif data_type == 'appointments':
            rows = conn.execute("""SELECT a.id, c.name, ca.plate, a.service, a.date, a.time, a.status, a.assigned_to
                FROM appointments a JOIN cars ca ON a.car_id=ca.id JOIN customers c ON ca.customer_id=c.id
                ORDER BY a.date DESC""").fetchall()
            headers = ['ID', 'Client', 'Plaque', 'Service', 'Date', 'Heure', 'Statut', 'Technicien']
        elif data_type == 'cars':
            rows = conn.execute("""SELECT ca.id, c.name, ca.brand, ca.model, ca.plate, ca.year, ca.color, ca.mileage
                FROM cars ca JOIN customers c ON ca.customer_id=c.id ORDER BY c.name""").fetchall()
            headers = ['ID', 'Propriétaire', 'Marque', 'Modèle', 'Plaque', 'Année', 'Couleur', 'Kilométrage']
        else:  # expenses
            rows = conn.execute("SELECT id, date, category, description, amount FROM expenses ORDER BY date DESC").fetchall()
            headers = ['ID', 'Date', 'Catégorie', 'Description', 'Montant']

    if fmt == 'csv':
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(headers)
        for row in rows:
            writer.writerow(row)
        resp = make_response(output.getvalue())
        resp.headers['Content-Type'] = 'text/csv; charset=utf-8'
        resp.headers['Content-Disposition'] = f'attachment; filename=amilcar_{data_type}.csv'
        return resp
    else:  # pdf
        try:
            from xhtml2pdf import pisa
        except ImportError:
            flash("xhtml2pdf non installé", "error")
            return redirect("/export_data")
        html = f"<html><head><meta charset='utf-8'><style>body{{font-family:Helvetica,sans-serif;font-size:11px}}table{{width:100%;border-collapse:collapse}}th,td{{border:1px solid #ccc;padding:5px;text-align:left}}th{{background:#D4AF37;color:#fff}}</style></head><body>"
        html += f"<h2>AMILCAR — {data_type.upper()}</h2><table><tr>"
        for h in headers:
            html += f"<th>{h}</th>"
        html += "</tr>"
        for row in rows:
            html += "<tr>" + "".join(f"<td>{c}</td>" for c in row) + "</tr>"
        html += "</table></body></html>"
        pdf_out = io.BytesIO()
        pisa.CreatePDF(io.StringIO(html), dest=pdf_out)
        resp = make_response(pdf_out.getvalue())
        resp.headers['Content-Type'] = 'application/pdf'
        resp.headers['Content-Disposition'] = f'attachment; filename=amilcar_{data_type}.pdf'
        return resp



@exports_bp.route('/export/customers_excel')
@login_required
def export_customers_excel():
    """Exporte la liste des clients en Excel (.xlsx)"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    import io

    with get_db() as conn:
        customers = conn.execute("""
            SELECT c.id, c.name, c.phone, c.email,
                   COALESCE(c.total_visits, 0) as visits,
                   COALESCE(c.last_visit, '') as last_visit,
                   COALESCE(rp.points, 0) as points,
                   COALESCE(rp.tier, 'Bronze') as tier,
                   COALESCE(SUM(i.amount), 0) as total_spent
            FROM customers c
            LEFT JOIN reward_points rp ON rp.customer_id = c.id
            LEFT JOIN cars ca ON ca.customer_id = c.id
            LEFT JOIN appointments a ON a.car_id = ca.id
            LEFT JOIN invoices i ON i.appointment_id = a.id AND i.status = 'paid'
            GROUP BY c.id ORDER BY total_spent DESC
        """).fetchall()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Clients"

    # Header style
    header_fill = PatternFill(start_color="1A1A2E", end_color="1A1A2E", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    border = Border(
        left=Side(style='thin', color='DDDDDD'),
        right=Side(style='thin', color='DDDDDD'),
        bottom=Side(style='thin', color='DDDDDD'),
    )

    headers = ["#", "Nom", "Téléphone", "Email", "Visites", "Dernière visite", "Points", "Tier", "CA Total (DT)"]
    col_widths = [5, 25, 15, 30, 8, 15, 8, 10, 15]

    # Title row
    ws.merge_cells("A1:I1")
    title_cell = ws["A1"]
    title_cell.value = "AMILCAR Auto Care — Liste des Clients"
    title_cell.font = Font(bold=True, size=14, color="1A1A2E")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    # Header row
    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[2].height = 22

    # Data rows
    tier_colors = {"Platinum": "E8C547", "Gold": "FFD700", "Silver": "C0C0C0", "Bronze": "CD7F32"}
    for row_idx, c in enumerate(customers, 3):
        row_fill = PatternFill(start_color="F8F9FA" if row_idx % 2 == 0 else "FFFFFF",
                               end_color="F8F9FA" if row_idx % 2 == 0 else "FFFFFF",
                               fill_type="solid")
        values = [c['id'], c['name'], c['phone'], c['email'] or '',
                  c['visits'], c['last_visit'], c['points'], c['tier'],
                  round(c['total_spent'], 2)]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.fill = row_fill
            cell.border = border
            cell.alignment = Alignment(vertical="center",
                                       horizontal="right" if col in (1, 5, 7, 9) else "left")
        # Color tier cell
        tier = c['tier']
        if tier in tier_colors:
            ws.cell(row=row_idx, column=8).fill = PatternFill(
                start_color=tier_colors[tier], end_color=tier_colors[tier], fill_type="solid")

    # Summary row
    last_row = len(customers) + 3
    ws.cell(row=last_row, column=1, value="TOTAL").font = Font(bold=True)
    ws.cell(row=last_row, column=5, value=f"=SUM(E3:E{last_row-1})").font = Font(bold=True)
    ws.cell(row=last_row, column=9, value=f"=SUM(I3:I{last_row-1})").font = Font(bold=True)

    # Freeze header
    ws.freeze_panes = "A3"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    log_activity('Export', 'Clients → Excel')
    return send_file(buf, as_attachment=True,
                     download_name=f"clients_amilcar_{__import__('datetime').date.today()}.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@exports_bp.route('/export/invoices_excel')
@login_required
def export_invoices_excel():
    """Exporte les factures en Excel"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    import io

    date_from = request.args.get('from', '')
    date_to = request.args.get('to', '')

    with get_db() as conn:
        q = """SELECT i.id, cu.name, a.date, a.service,
                      i.amount, i.paid_amount, i.status, i.payment_method,
                      ca.brand || ' ' || ca.model as car, ca.plate
               FROM invoices i
               JOIN appointments a ON i.appointment_id = a.id
               JOIN cars ca ON a.car_id = ca.id
               JOIN customers cu ON ca.customer_id = cu.id
               WHERE 1=1"""
        params = []
        if date_from:
            q += " AND a.date >= ?"
            params.append(date_from)
        if date_to:
            q += " AND a.date <= ?"
            params.append(date_to)
        q += " ORDER BY i.id DESC"
        invoices = conn.execute(q, params).fetchall()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Factures"

    header_fill = PatternFill(start_color="1A1A2E", end_color="1A1A2E", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    border = Border(
        left=Side(style='thin', color='DDDDDD'),
        right=Side(style='thin', color='DDDDDD'),
        bottom=Side(style='thin', color='DDDDDD'),
    )

    headers = ["N°", "Client", "Date", "Service", "Voiture", "Plaque", "Montant DT", "Payé DT", "Reste DT", "Statut", "Paiement"]
    col_widths = [6, 22, 12, 28, 20, 12, 12, 12, 12, 10, 12]

    ws.merge_cells("A1:K1")
    ws["A1"].value = f"AMILCAR — Factures{'  ' + date_from + ' → ' + date_to if date_from else ''}"
    ws["A1"].font = Font(bold=True, size=13, color="1A1A2E")
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 28

    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = border
        ws.column_dimensions[get_column_letter(col)].width = w

    status_colors = {"paid": "C8E6C9", "unpaid": "FFCDD2", "partial": "FFF9C4", "cancelled": "EEEEEE"}

    total_amount = total_paid = 0
    for row_idx, inv in enumerate(invoices, 3):
        amount = inv['amount'] or 0
        paid = inv['paid_amount'] or 0
        reste = round(amount - paid, 2)
        total_amount += amount
        total_paid += paid
        row_fill = PatternFill(start_color="F8F9FA" if row_idx % 2 == 0 else "FFFFFF",
                               end_color="F8F9FA" if row_idx % 2 == 0 else "FFFFFF",
                               fill_type="solid")
        values = [inv['id'], inv['name'], inv['date'], inv['service'],
                  inv['car'], inv['plate'], round(amount, 2), round(paid, 2), reste,
                  inv['status'], inv['payment_method'] or '']
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.fill = row_fill
            cell.border = border
            cell.alignment = Alignment(vertical="center",
                                       horizontal="right" if col in (7, 8, 9) else "left")
        # Color status
        st = inv['status']
        if st in status_colors:
            ws.cell(row=row_idx, column=10).fill = PatternFill(
                start_color=status_colors[st], end_color=status_colors[st], fill_type="solid")

    # Totals
    last = len(invoices) + 3
    for col, val in [(1, "TOTAL"), (7, round(total_amount, 2)),
                     (8, round(total_paid, 2)), (9, round(total_amount - total_paid, 2))]:
        cell = ws.cell(row=last, column=col, value=val)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")

    ws.freeze_panes = "A3"
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    log_activity('Export', 'Factures → Excel')
    return send_file(buf, as_attachment=True,
                     download_name=f"factures_amilcar_{__import__('datetime').date.today()}.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@exports_bp.route('/export/appointments_excel')
@login_required
def export_appointments_excel():
    """Exporte les rendez-vous en Excel"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    import io

    date_from = request.args.get('from', '')
    date_to = request.args.get('to', '')

    with get_db() as conn:
        q = """SELECT a.id, cu.name, cu.phone, ca.brand || ' ' || ca.model as car,
                      ca.plate, a.date, a.time, a.service, a.status
               FROM appointments a
               JOIN cars ca ON a.car_id = ca.id
               JOIN customers cu ON ca.customer_id = cu.id
               WHERE 1=1"""
        params = []
        if date_from:
            q += " AND a.date >= ?"
            params.append(date_from)
        if date_to:
            q += " AND a.date <= ?"
            params.append(date_to)
        q += " ORDER BY a.date DESC, a.time"
        appts = conn.execute(q, params).fetchall()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Rendez-vous"

    header_fill = PatternFill(start_color="0F3460", end_color="0F3460", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    border = Border(
        left=Side(style='thin', color='DDDDDD'),
        right=Side(style='thin', color='DDDDDD'),
        bottom=Side(style='thin', color='DDDDDD'),
    )

    headers = ["#", "Client", "Téléphone", "Véhicule", "Plaque", "Date", "Heure", "Service", "Statut"]
    col_widths = [5, 22, 14, 20, 12, 12, 8, 28, 12]

    ws.merge_cells("A1:I1")
    ws["A1"].value = "AMILCAR Auto Care — Rendez-vous"
    ws["A1"].font = Font(bold=True, size=13, color="0F3460")
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 28

    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = border
        ws.column_dimensions[get_column_letter(col)].width = w

    status_colors = {
        "completed": "C8E6C9", "pending": "FFF9C4",
        "cancelled": "FFCDD2", "in_progress": "BBDEFB"
    }
    for row_idx, a in enumerate(appts, 3):
        row_fill = PatternFill(start_color="F8F9FA" if row_idx % 2 == 0 else "FFFFFF",
                               end_color="F8F9FA" if row_idx % 2 == 0 else "FFFFFF",
                               fill_type="solid")
        values = [a['id'], a['name'], a['phone'], a['car'], a['plate'],
                  a['date'], a['time'] or '', a['service'], a['status']]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.fill = row_fill
            cell.border = border
            cell.alignment = Alignment(vertical="center")
        st = a['status']
        if st in status_colors:
            ws.cell(row=row_idx, column=9).fill = PatternFill(
                start_color=status_colors[st], end_color=status_colors[st], fill_type="solid")

    ws.freeze_panes = "A3"
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    log_activity('Export', 'RDV → Excel')
    return send_file(buf, as_attachment=True,
                     download_name=f"rdv_amilcar_{__import__('datetime').date.today()}.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@exports_bp.route('/export/full_report_excel')
@exports_bp.route('/export/monthly')
@login_required
def export_monthly_report_excel():
    """Rapport mensuel complet en Excel avec plusieurs feuilles"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.chart import BarChart, Reference
    from openpyxl.utils import get_column_letter
    import io
    from datetime import date

    month = request.args.get('month', date.today().strftime('%Y-%m'))

    with get_db() as conn:
        invoices = conn.execute("""
            SELECT i.id, cu.name, a.date, a.service, i.amount, i.status, i.payment_method
            FROM invoices i JOIN appointments a ON i.appointment_id=a.id
            JOIN cars ca ON a.car_id=ca.id JOIN customers cu ON ca.customer_id=cu.id
            WHERE strftime('%Y-%m', a.date) = ? ORDER BY a.date
        """, (month,)).fetchall()

        expenses = conn.execute("""
            SELECT id, description, amount, category, date
            FROM expenses WHERE strftime('%Y-%m', date) = ? ORDER BY date
        """, (month,)).fetchall()

        top_services = conn.execute("""
            SELECT a.service, COUNT(*) as cnt, COALESCE(SUM(i.amount),0) as revenue
            FROM appointments a LEFT JOIN invoices i ON a.id=i.appointment_id
            WHERE strftime('%Y-%m', a.date) = ? AND a.status='completed'
            GROUP BY a.service ORDER BY revenue DESC LIMIT 10
        """, (month,)).fetchall()

        new_customers = conn.execute(
            "SELECT COUNT(*) FROM customers WHERE strftime('%Y-%m', last_visit) = ?",
            (month,)).fetchone()[0]

    wb = openpyxl.Workbook()
    header_fill = PatternFill(start_color="1A1A2E", end_color="1A1A2E", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    border = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC'),
    )

    def style_header(ws, row, cols, headers, widths):
        for col, (h, w) in enumerate(zip(headers, widths), 1):
            cell = ws.cell(row=row, column=col, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
            cell.border = border
            ws.column_dimensions[get_column_letter(col)].width = w

    # ── Feuille 1: Résumé ──────────────────────────────────────────────────
    ws_sum = wb.active
    ws_sum.title = "Résumé"
    total_revenue = sum(r['amount'] for r in invoices if r['status'] == 'paid')
    total_invoices = len(invoices)
    total_expenses = sum(e['amount'] for e in expenses)
    profit = total_revenue - total_expenses

    summary_data = [
        ("Mois", month),
        ("Chiffre d'affaires (DT)", round(total_revenue, 2)),
        ("Nombre de factures", total_invoices),
        ("Dépenses totales (DT)", round(total_expenses, 2)),
        ("Bénéfice net (DT)", round(profit, 2)),
        ("Nouveaux clients", new_customers),
        ("Taux de marge (%)", round((profit / total_revenue * 100) if total_revenue else 0, 1)),
    ]
    ws_sum.merge_cells("A1:B1")
    ws_sum["A1"].value = f"AMILCAR — Rapport Mensuel {month}"
    ws_sum["A1"].font = Font(bold=True, size=14, color="1A1A2E")
    ws_sum["A1"].alignment = Alignment(horizontal="center")
    ws_sum.row_dimensions[1].height = 30
    ws_sum.column_dimensions["A"].width = 30
    ws_sum.column_dimensions["B"].width = 20

    for r, (label, val) in enumerate(summary_data, 2):
        ws_sum.cell(row=r, column=1, value=label).font = Font(bold=True)
        cell = ws_sum.cell(row=r, column=2, value=val)
        if "Bénéfice" in label:
            cell.font = Font(bold=True, color="2E7D32" if profit >= 0 else "C62828")
        ws_sum.row_dimensions[r].height = 20

    # ── Feuille 2: Factures ────────────────────────────────────────────────
    ws_inv = wb.create_sheet("Factures")
    ws_inv.merge_cells("A1:G1")
    ws_inv["A1"].value = f"Factures — {month}"
    ws_inv["A1"].font = Font(bold=True, size=12, color="1A1A2E")
    ws_inv["A1"].alignment = Alignment(horizontal="center")
    style_header(ws_inv, 2,
                 ["#", "Client", "Date", "Service", "Montant DT", "Statut", "Paiement"],
                 ["#", "Client", "Date", "Service", "Montant DT", "Statut", "Paiement"],
                 [5, 22, 12, 28, 12, 10, 12])
    for ri, inv in enumerate(invoices, 3):
        row_fill = PatternFill(start_color="F8F9FA" if ri % 2 == 0 else "FFFFFF",
                               end_color="F8F9FA" if ri % 2 == 0 else "FFFFFF", fill_type="solid")
        for ci, val in enumerate([inv['id'], inv['name'], inv['date'], inv['service'],
                                   round(inv['amount'], 2), inv['status'], inv['payment_method'] or ''], 1):
            c = ws_inv.cell(row=ri, column=ci, value=val)
            c.fill = row_fill
            c.border = border

    # ── Feuille 3: Top Services ────────────────────────────────────────────
    ws_svc = wb.create_sheet("Top Services")
    ws_svc.merge_cells("A1:C1")
    ws_svc["A1"].value = f"Top Services — {month}"
    ws_svc["A1"].font = Font(bold=True, size=12, color="1A1A2E")
    ws_svc["A1"].alignment = Alignment(horizontal="center")
    style_header(ws_svc, 2,
                 ["Service", "Nombre", "CA (DT)"],
                 ["Service", "Nombre", "CA (DT)"],
                 [30, 10, 14])
    for ri, svc in enumerate(top_services, 3):
        for ci, val in enumerate([svc['service'], svc['cnt'], round(svc['revenue'], 2)], 1):
            c = ws_svc.cell(row=ri, column=ci, value=val)
            c.border = border

    # ── Feuille 4: Dépenses ────────────────────────────────────────────────
    ws_exp = wb.create_sheet("Dépenses")
    ws_exp.merge_cells("A1:E1")
    ws_exp["A1"].value = f"Dépenses — {month}"
    ws_exp["A1"].font = Font(bold=True, size=12, color="1A1A2E")
    ws_exp["A1"].alignment = Alignment(horizontal="center")
    style_header(ws_exp, 2,
                 ["#", "Description", "Montant DT", "Catégorie", "Date"],
                 ["#", "Description", "Montant DT", "Catégorie", "Date"],
                 [5, 30, 14, 16, 12])
    for ri, exp in enumerate(expenses, 3):
        for ci, val in enumerate([exp['id'], exp['description'],
                                   round(exp['amount'], 2), exp['category'], exp['date']], 1):
            c = ws_exp.cell(row=ri, column=ci, value=val)
            c.border = border

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    log_activity('Export', f'Rapport mensuel {month} → Excel')
    return send_file(buf, as_attachment=True,
                     download_name=f"rapport_{month}_amilcar.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@exports_bp.route('/export/inventory_excel')
@login_required
def export_inventory_excel():
    """Exporte l'inventaire en Excel"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    import io

    with get_db() as conn:
        items = conn.execute(
            "SELECT id, name, category, quantity, min_quantity, unit_price FROM inventory ORDER BY category, name"
        ).fetchall()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Inventaire"

    header_fill = PatternFill(start_color="0F3460", end_color="0F3460", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    border = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC'),
    )

    ws.merge_cells("A1:F1")
    ws["A1"].value = f"AMILCAR — Inventaire au {__import__('datetime').date.today()}"
    ws["A1"].font = Font(bold=True, size=13, color="0F3460")
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 28

    headers = ["#", "Article", "Catégorie", "Quantité", "Qté Min", "Prix Unitaire DT"]
    col_widths = [5, 30, 16, 10, 10, 18]
    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = border
        ws.column_dimensions[get_column_letter(col)].width = w

    low_stock_fill = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")
    ok_stock_fill  = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")

    for row_idx, item in enumerate(items, 3):
        is_low = (item['quantity'] or 0) <= (item['min_quantity'] or 0)
        row_fill = low_stock_fill if is_low else (
            PatternFill(start_color="F8F9FA" if row_idx % 2 == 0 else "FFFFFF",
                        end_color="F8F9FA" if row_idx % 2 == 0 else "FFFFFF",
                        fill_type="solid"))
        values = [item['id'], item['name'], item['category'],
                  item['quantity'], item['min_quantity'],
                  round(item['unit_price'] or 0, 2)]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.fill = row_fill
            cell.border = border
            cell.alignment = Alignment(vertical="center",
                                       horizontal="right" if col in (4, 5, 6) else "left")

    # Legend
    last = len(items) + 4
    ws.cell(row=last, column=1, value="🔴 Rouge = stock bas").font = Font(italic=True, color="C62828")
    ws.cell(row=last+1, column=1, value="🟢 Vert = stock OK").font = Font(italic=True, color="2E7D32")

    ws.freeze_panes = "A3"
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    log_activity('Export', 'Inventaire → Excel')
    return send_file(buf, as_attachment=True,
                     download_name=f"inventaire_amilcar_{__import__('datetime').date.today()}.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


# ── PDF Exports ──────────────────────────────────────────────────────────────

def _render_pdf(html_content):
    """Helper: convert HTML to PDF bytes"""
    from xhtml2pdf import pisa
    import io
    buf = io.BytesIO()
    pisa.CreatePDF(io.StringIO(html_content), dest=buf)
    buf.seek(0)
    return buf


@exports_bp.route('/export/monthly_pdf')
@login_required
def export_monthly_pdf():
    """Rapport mensuel en PDF"""
    from datetime import date as dt_date
    month = request.args.get('month', dt_date.today().strftime('%Y-%m'))

    with get_db() as conn:
        invoices = conn.execute("""
            SELECT i.id, cu.name, a.date, a.service, i.amount, i.status, i.payment_method
            FROM invoices i JOIN appointments a ON i.appointment_id=a.id
            JOIN cars ca ON a.car_id=ca.id JOIN customers cu ON ca.customer_id=cu.id
            WHERE strftime('%Y-%m', a.date) = ? ORDER BY a.date
        """, (month,)).fetchall()

        expenses = conn.execute(
            "SELECT description, amount, category FROM expenses WHERE strftime('%Y-%m', date) = ? ORDER BY date",
            (month,)).fetchall()

        top_services = conn.execute("""
            SELECT a.service, COUNT(*) as cnt, COALESCE(SUM(i.amount),0) as revenue
            FROM appointments a LEFT JOIN invoices i ON a.id=i.appointment_id
            WHERE strftime('%Y-%m', a.date) = ? AND a.status='completed'
            GROUP BY a.service ORDER BY revenue DESC LIMIT 8
        """, (month,)).fetchall()

        settings = get_all_settings()

    total_revenue = sum(r['amount'] for r in invoices if r['status'] == 'paid')
    total_expenses_sum = sum(e['amount'] for e in expenses)
    profit = total_revenue - total_expenses_sum
    shop_name = settings.get('shop_name', 'AMILCAR Auto Care')

    profit_color = '#2e7d32' if profit >= 0 else '#c62828'
    rows_invoices = ''.join(
        f"<tr><td>{r['id']}</td><td>{r['name']}</td><td>{r['date']}</td>"
        f"<td>{r['service']}</td><td>{round(r['amount'],2)} DT</td>"
        f"<td class='{r['status']}'>{r['status']}</td></tr>"
        for r in invoices[:50]
    )
    rows_services = ''.join(
        f"<tr><td>{s['service']}</td><td>{s['cnt']}</td><td>{round(s['revenue'],2)}</td></tr>"
        for s in top_services
    )
    rows_expenses = ''.join(
        f"<tr><td>{e['description']}</td><td>{round(e['amount'],2)} DT</td><td>{e['category']}</td></tr>"
        for e in expenses
    )

    html = f"""<!DOCTYPE html>
<html><head><meta charset="utf-8">
<style>
  body {{ font-family: Arial, sans-serif; font-size: 11px; color: #1a1a2e; margin: 20px; }}
  h1 {{ color: #1a1a2e; font-size: 20px; text-align: center; border-bottom: 2px solid #e8c547; padding-bottom: 8px; }}
  h2 {{ color: #0f3460; font-size: 14px; margin-top: 20px; border-left: 4px solid #e8c547; padding-left: 8px; }}
  .kpi-grid {{ display: table; width: 100%; margin: 16px 0; }}
  .kpi {{ display: table-cell; text-align: center; padding: 10px; background: #f8f9fa; border: 1px solid #ddd; }}
  .kpi-val {{ font-size: 20px; font-weight: bold; color: #0f3460; }}
  .kpi-label {{ font-size: 10px; color: #666; }}
  table {{ width: 100%; border-collapse: collapse; margin-top: 8px; }}
  th {{ background: #1a1a2e; color: white; padding: 6px 8px; text-align: left; font-size: 10px; }}
  td {{ padding: 5px 8px; border-bottom: 1px solid #eee; font-size: 10px; }}
  tr:nth-child(even) td {{ background: #f8f9fa; }}
  .paid {{ color: #2e7d32; font-weight: bold; }}
  .unpaid {{ color: #c62828; font-weight: bold; }}
  .partial {{ color: #e65100; font-weight: bold; }}
  .footer {{ text-align: center; color: #999; font-size: 9px; margin-top: 30px; border-top: 1px solid #eee; padding-top: 8px; }}
</style>
</head><body>
<h1>AMILCAR — Rapport Mensuel {month}</h1>
<div class="kpi-grid">
  <div class="kpi"><div class="kpi-val">{round(total_revenue, 0):.0f} DT</div><div class="kpi-label">Chiffre d&apos;affaires</div></div>
  <div class="kpi"><div class="kpi-val">{len(invoices)}</div><div class="kpi-label">Factures</div></div>
  <div class="kpi"><div class="kpi-val">{round(total_expenses_sum, 0):.0f} DT</div><div class="kpi-label">Depenses</div></div>
  <div class="kpi"><div class="kpi-val" style="color:{profit_color}">{round(profit, 0):.0f} DT</div><div class="kpi-label">Benefice net</div></div>
</div>
<h2>Top Services</h2>
<table><tr><th>Service</th><th>Nombre</th><th>Revenu (DT)</th></tr>{rows_services}</table>
<h2>Factures ({len(invoices)})</h2>
<table><tr><th>#</th><th>Client</th><th>Date</th><th>Service</th><th>Montant</th><th>Statut</th></tr>{rows_invoices}</table>
<h2>Depenses ({len(expenses)})</h2>
<table><tr><th>Description</th><th>Montant</th><th>Categorie</th></tr>{rows_expenses}</table>
<div class="footer">Genere le {dt_date.today()} — {shop_name}</div>
</body></html>"""

    buf = _render_pdf(html)
    log_activity('Export', f'Rapport {month} → PDF')
    return send_file(buf, as_attachment=True,
                     download_name=f"rapport_{month}_amilcar.pdf",
                     mimetype="application/pdf")


@exports_bp.route('/export/professional_pdf')
@login_required
def export_professional_pdf():
    """Rapport general professionnel PDF"""
    from datetime import date as dt_date

    today = dt_date.today()
    month_start = today.strftime('%Y-%m-01')

    with get_db() as conn:
        settings = get_all_settings()
        total_customers = conn.execute("SELECT COUNT(*) FROM customers").fetchone()[0]
        total_cars = conn.execute("SELECT COUNT(*) FROM cars").fetchone()[0]
        revenue_month = conn.execute(
            "SELECT COALESCE(SUM(amount),0) FROM invoices WHERE status='paid' AND created_at >= ?",
            (month_start,)).fetchone()[0]
        revenue_total = conn.execute(
            "SELECT COALESCE(SUM(amount),0) FROM invoices WHERE status='paid'").fetchone()[0]
        appts_total = conn.execute("SELECT COUNT(*) FROM appointments").fetchone()[0]
        appts_completed = conn.execute(
            "SELECT COUNT(*) FROM appointments WHERE status='completed'").fetchone()[0]
        top_services = conn.execute("""
            SELECT a.service, COUNT(*) as cnt, COALESCE(SUM(i.amount),0) as rev
            FROM appointments a LEFT JOIN invoices i ON a.id=i.appointment_id
            WHERE a.status='completed' GROUP BY a.service ORDER BY rev DESC LIMIT 8
        """).fetchall()
        recent_invoices = conn.execute("""
            SELECT i.id, cu.name, a.date, a.service, i.amount, i.status
            FROM invoices i JOIN appointments a ON i.appointment_id=a.id
            JOIN cars ca ON a.car_id=ca.id JOIN customers cu ON ca.customer_id=cu.id
            ORDER BY i.id DESC LIMIT 15
        """).fetchall()
        low_stock = conn.execute(
            "SELECT name, quantity, min_quantity FROM inventory WHERE quantity <= min_quantity AND min_quantity > 0"
        ).fetchall()

    shop_name = settings.get('shop_name', 'AMILCAR Auto Care')
    completion_rate = round(appts_completed / appts_total * 100, 1) if appts_total else 0

    rows_services = ''.join(
        f"<tr><td>{s['service']}</td><td>{s['cnt']}</td><td>{round(s['rev'],2)}</td></tr>"
        for s in top_services
    )
    rows_invoices = ''.join(
        f"<tr><td>{r['id']}</td><td>{r['name']}</td><td>{r['date']}</td>"
        f"<td>{r['service']}</td><td>{round(r['amount'],2)}</td>"
        f"<td class='badge-{r['status']}'>{r['status']}</td></tr>"
        for r in recent_invoices
    )
    stock_alerts = ''.join(
        f"<p>- {i['name']} : {i['quantity']} unites (min: {i['min_quantity']})</p>"
        for i in low_stock
    )
    stock_section = (
        f'<h2>Alertes Stock</h2><div class="alert-box">{stock_alerts}</div>'
        if low_stock else ''
    )

    html = f"""<!DOCTYPE html>
<html><head><meta charset="utf-8">
<style>
  body {{ font-family: Arial, sans-serif; font-size: 11px; color: #1a1a2e; margin: 24px; }}
  .header {{ background: #1a1a2e; color: white; padding: 16px; text-align: center; }}
  .header h1 {{ color: #e8c547; margin: 0; font-size: 20px; }}
  .header p {{ color: #aaa; margin: 4px 0 0; }}
  h2 {{ color: #0f3460; font-size: 13px; margin-top: 22px; border-bottom: 2px solid #e8c547; padding-bottom: 4px; }}
  .kpi-row {{ display: table; width: 100%; margin: 12px 0; }}
  .kpi {{ display: table-cell; padding: 10px; background: #f0f4ff; border: 1px solid #c5cef0; text-align: center; }}
  .kpi-val {{ font-size: 20px; font-weight: bold; color: #0f3460; }}
  table {{ width: 100%; border-collapse: collapse; margin-top: 8px; font-size: 10px; }}
  th {{ background: #0f3460; color: white; padding: 6px 8px; text-align: left; }}
  td {{ padding: 5px 8px; border-bottom: 1px solid #eee; }}
  tr:nth-child(even) td {{ background: #f8f9fa; }}
  .badge-paid {{ color: #2e7d32; font-weight: bold; }}
  .badge-unpaid {{ color: #c62828; }}
  .badge-partial {{ color: #e65100; }}
  .alert-box {{ background: #fff3cd; border: 1px solid #ffc107; padding: 10px; margin: 10px 0; }}
  .footer {{ text-align: center; color: #aaa; font-size: 9px; margin-top: 30px; border-top: 1px solid #ddd; padding-top: 8px; }}
</style>
</head><body>
<div class="header">
  <h1>{shop_name}</h1>
  <p>Rapport Professionnel — {today.strftime('%d/%m/%Y')}</p>
</div>
<h2>Indicateurs Cles</h2>
<div class="kpi-row">
  <div class="kpi"><div class="kpi-val">{total_customers}</div><div>Clients</div></div>
  <div class="kpi"><div class="kpi-val">{total_cars}</div><div>Vehicules</div></div>
  <div class="kpi"><div class="kpi-val">{appts_total}</div><div>RDV total</div></div>
  <div class="kpi"><div class="kpi-val">{completion_rate}%</div><div>Taux completion</div></div>
</div>
<div class="kpi-row">
  <div class="kpi"><div class="kpi-val">{round(revenue_month):,} DT</div><div>CA ce mois</div></div>
  <div class="kpi"><div class="kpi-val">{round(revenue_total):,} DT</div><div>CA total</div></div>
</div>
<h2>Top Services</h2>
<table><tr><th>Service</th><th>Prestations</th><th>Revenu (DT)</th></tr>{rows_services}</table>
<h2>Dernieres Factures</h2>
<table><tr><th>#</th><th>Client</th><th>Date</th><th>Service</th><th>Montant DT</th><th>Statut</th></tr>{rows_invoices}</table>
{stock_section}
<div class="footer">{shop_name} — Rapport confidentiel — {today}</div>
</body></html>"""

    buf = _render_pdf(html)
    log_activity('Export', 'Rapport professionnel → PDF')
    return send_file(buf, as_attachment=True,
                     download_name=f"rapport_professionnel_{today}.pdf",
                     mimetype="application/pdf")


@exports_bp.route('/export/daily')
@login_required
def export_daily_pdf():
    """Rapport journalier PDF"""
    from datetime import date as dt_date

    day = request.args.get('date', dt_date.today().isoformat())

    with get_db() as conn:
        settings = get_all_settings()
        appts = conn.execute("""
            SELECT a.id, cu.name, cu.phone, ca.brand || ' ' || ca.model as car,
                   ca.plate, a.time, a.service, a.status,
                   COALESCE(i.amount, 0) as amount
            FROM appointments a
            JOIN cars ca ON a.car_id = ca.id
            JOIN customers cu ON ca.customer_id = cu.id
            LEFT JOIN invoices i ON i.appointment_id = a.id
            WHERE a.date = ? ORDER BY a.time
        """, (day,)).fetchall()

        revenue = conn.execute(
            "SELECT COALESCE(SUM(i.amount),0) FROM invoices i "
            "JOIN appointments a ON i.appointment_id=a.id WHERE a.date=? AND i.status='paid'",
            (day,)).fetchone()[0]

    shop_name = settings.get('shop_name', 'AMILCAR Auto Care')
    completed = sum(1 for a in appts if a['status'] == 'completed')

    rows = ''.join(
        f"<tr><td>{a['time'] or '-'}</td><td>{a['name']}</td><td>{a['car']}</td>"
        f"<td>{a['plate']}</td><td>{a['service']}</td>"
        f"<td class='{a['status']}'>{a['status']}</td>"
        f"<td>{round(a['amount'],2) if a['amount'] else '-'} DT</td></tr>"
        for a in appts
    )

    html = f"""<!DOCTYPE html>
<html><head><meta charset="utf-8">
<style>
  body {{ font-family: Arial, sans-serif; font-size: 11px; color: #1a1a2e; margin: 24px; }}
  .header {{ background: #1a1a2e; color: white; padding: 14px; text-align: center; }}
  .header h1 {{ color: #e8c547; margin: 0; font-size: 18px; }}
  .stats {{ display: table; width: 100%; margin: 14px 0; }}
  .stat {{ display: table-cell; text-align: center; padding: 10px; background: #f0f4ff; border: 1px solid #ddd; }}
  .stat-val {{ font-size: 20px; font-weight: bold; color: #0f3460; }}
  table {{ width: 100%; border-collapse: collapse; margin-top: 12px; }}
  th {{ background: #0f3460; color: white; padding: 7px; font-size: 10px; }}
  td {{ padding: 6px 7px; border-bottom: 1px solid #eee; font-size: 10px; }}
  tr:nth-child(even) td {{ background: #f8f9fa; }}
  .completed {{ color: #2e7d32; font-weight: bold; }}
  .pending {{ color: #e65100; }}
  .cancelled {{ color: #c62828; }}
</style>
</head><body>
<div class="header">
  <h1>{shop_name} — Rapport Journalier</h1>
  <p style="color:#aaa;margin:4px 0 0">{day}</p>
</div>
<div class="stats">
  <div class="stat"><div class="stat-val">{len(appts)}</div><div>RDV total</div></div>
  <div class="stat"><div class="stat-val">{completed}</div><div>Completes</div></div>
  <div class="stat"><div class="stat-val">{round(revenue,0):.0f} DT</div><div>Revenu</div></div>
</div>
<table>
  <tr><th>Heure</th><th>Client</th><th>Vehicule</th><th>Plaque</th><th>Service</th><th>Statut</th><th>Montant</th></tr>
  {rows}
</table>
</body></html>"""

    buf = _render_pdf(html)
    log_activity('Export', f'Rapport journalier {day} → PDF')
    return send_file(buf, as_attachment=True,
                     download_name=f"rapport_journalier_{day}.pdf",
                     mimetype="application/pdf")


@exports_bp.route('/export/expenses')
@login_required
def export_expenses_excel():
    """Exporte les depenses en Excel"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    import io
    from datetime import date as dt_date

    month = request.args.get('month', dt_date.today().strftime('%Y-%m'))

    with get_db() as conn:
        expenses = conn.execute(
            "SELECT id, description, amount, category, date FROM expenses "
            "WHERE strftime('%Y-%m', date) = ? ORDER BY date",
            (month,)).fetchall()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Depenses"

    header_fill = PatternFill(start_color="C62828", end_color="C62828", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    border = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC'),
    )

    ws.merge_cells("A1:E1")
    ws["A1"].value = f"AMILCAR — Depenses {month}"
    ws["A1"].font = Font(bold=True, size=13, color="C62828")
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 28

    headers = ["#", "Description", "Montant DT", "Categorie", "Date"]
    col_widths = [5, 35, 14, 18, 12]
    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = border
        ws.column_dimensions[get_column_letter(col)].width = w

    total = 0
    for row_idx, exp in enumerate(expenses, 3):
        total += exp['amount'] or 0
        row_fill = PatternFill(start_color="F8F9FA" if row_idx % 2 == 0 else "FFFFFF",
                               end_color="F8F9FA" if row_idx % 2 == 0 else "FFFFFF",
                               fill_type="solid")
        for col, val in enumerate([exp['id'], exp['description'],
                                    round(exp['amount'], 2), exp['category'], exp['date']], 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.fill = row_fill
            cell.border = border
            cell.alignment = Alignment(horizontal="right" if col == 3 else "left")

    last_row = len(expenses) + 3
    ws.cell(row=last_row, column=2, value="TOTAL").font = Font(bold=True)
    total_cell = ws.cell(row=last_row, column=3, value=round(total, 2))
    total_cell.font = Font(bold=True, color="C62828")
    total_cell.fill = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")

    ws.freeze_panes = "A3"
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    log_activity('Export', f'Depenses {month} → Excel')
    return send_file(buf, as_attachment=True,
                     download_name=f"depenses_{month}_amilcar.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


# ─── Customer Statement PDF ───
@exports_bp.route('/export/customer_statement/<int:customer_id>')
@login_required
def export_customer_statement(customer_id):
    """Relevé de compte client PDF"""
    with get_db() as conn:
        cust = conn.execute("SELECT id, name, phone, email, address FROM customers WHERE id=?", (customer_id,)).fetchone()
        if not cust:
            flash("Client introuvable", "error")
            return redirect("/customers")
        invoices = conn.execute("""
            SELECT i.id, a.date, a.service, ca.brand || ' ' || ca.model as car, ca.plate,
                   i.amount, i.status, COALESCE(i.paid_amount, 0), i.payment_method
            FROM invoices i JOIN appointments a ON i.appointment_id=a.id
            JOIN cars ca ON a.car_id=ca.id
            WHERE ca.customer_id=? ORDER BY a.date DESC
        """, (customer_id,)).fetchall()
        cars = conn.execute("SELECT brand, model, plate FROM cars WHERE customer_id=?", (customer_id,)).fetchall()

    settings = get_all_settings()
    shop_name = settings.get('shop_name', 'AMILCAR Auto Care')
    today = date.today()

    total_billed = sum(float(inv['amount']) for inv in invoices)
    total_paid = sum(float(inv['paid_amount']) for inv in invoices)
    balance = total_billed - total_paid

    rows = ''.join(
        f"<tr><td>#{inv['id']}</td><td>{inv['date']}</td><td>{inv['service']}</td>"
        f"<td>{inv['car']}</td><td class='ar'>{round(inv['amount'],2)}</td>"
        f"<td class='ar'>{round(inv['paid_amount'],2)}</td>"
        f"<td class='st-{inv['status']}'>{inv['status']}</td></tr>"
        for inv in invoices
    )
    car_list = ', '.join(f"{c['brand']} {c['model']} ({c['plate']})" for c in cars)

    html = f"""<!DOCTYPE html>
<html><head><meta charset="utf-8">
<style>
  body {{ font-family: Arial, sans-serif; font-size: 11px; color: #1a1a2e; margin: 24px; }}
  .header {{ background: #1a1a2e; color: white; padding: 18px; display: flex; justify-content: space-between; }}
  .header h1 {{ color: #e8c547; margin: 0; font-size: 18px; letter-spacing: 3px; }}
  .header p {{ color: #aaa; margin: 4px 0 0; font-size: 9px; }}
  .title {{ text-align: center; color: #B8962E; font-size: 14px; letter-spacing: 4px; margin: 20px 0; font-weight: 700; }}
  .info {{ display: table; width: 100%; margin: 14px 0; }}
  .info-col {{ display: table-cell; width: 50%; padding: 10px; vertical-align: top; }}
  .info-col h3 {{ color: #B8962E; font-size: 10px; letter-spacing: 2px; margin: 0 0 6px; }}
  .info-col p {{ margin: 2px 0; font-size: 11px; }}
  .summary {{ display: table; width: 100%; margin: 16px 0; }}
  .sum-box {{ display: table-cell; text-align: center; padding: 12px; border: 1px solid #e0e0e0; }}
  .sum-val {{ font-size: 18px; font-weight: 800; }}
  .sum-green {{ color: #2e7d32; }}
  .sum-red {{ color: #c62828; }}
  .sum-gold {{ color: #B8962E; }}
  table {{ width: 100%; border-collapse: collapse; margin: 12px 0; font-size: 10px; }}
  th {{ background: #1a1a2e; color: #e8c547; padding: 7px 8px; text-align: left; font-size: 9px; letter-spacing: 1px; }}
  td {{ padding: 6px 8px; border-bottom: 1px solid #eee; }}
  tr:nth-child(even) td {{ background: #f8f9fa; }}
  .ar {{ text-align: right; }}
  .st-paid {{ color: #2e7d32; font-weight: 700; }}
  .st-unpaid {{ color: #c62828; font-weight: 700; }}
  .st-partial {{ color: #e65100; font-weight: 700; }}
  .footer {{ text-align: center; color: #aaa; font-size: 8px; margin-top: 30px; border-top: 1px solid #ddd; padding-top: 8px; }}
</style></head><body>
<div class="header"><div><h1>{shop_name}</h1><p>Relevé de Compte Client</p></div><div style="text-align:right"><p>Date: {today.strftime('%d/%m/%Y')}</p></div></div>
<div class="title">RELEVÉ DE COMPTE</div>
<div class="info">
  <div class="info-col"><h3>CLIENT</h3><p><b>{cust['name']}</b></p><p>{cust['phone'] or ''}</p><p>{cust['email'] or ''}</p><p>{cust['address'] or ''}</p></div>
  <div class="info-col"><h3>VÉHICULES</h3><p>{car_list or '—'}</p><p style="margin-top:8px;color:#888">Total factures: {len(invoices)}</p></div>
</div>
<div class="summary">
  <div class="sum-box"><div class="sum-val sum-gold">{round(total_billed,2)} DT</div><div>Total Facturé</div></div>
  <div class="sum-box"><div class="sum-val sum-green">{round(total_paid,2)} DT</div><div>Total Payé</div></div>
  <div class="sum-box"><div class="sum-val sum-red">{round(balance,2)} DT</div><div>Solde Restant</div></div>
</div>
<table><tr><th>#</th><th>Date</th><th>Service</th><th>Véhicule</th><th>Montant</th><th>Payé</th><th>Statut</th></tr>{rows}</table>
<div class="footer">{shop_name} — Document confidentiel — Généré le {today.strftime('%d/%m/%Y à %H:%M')}</div>
</body></html>"""

    buf = _render_pdf(html)
    log_activity('Export', f'Relevé client #{customer_id} → PDF')
    safe_name = re.sub(r'[^a-zA-Z0-9]', '_', cust['name'])
    return send_file(buf, as_attachment=True,
                     download_name=f"releve_{safe_name}_{today}.pdf",
                     mimetype="application/pdf")


# ─── Quote PDF ───
@exports_bp.route('/export/quote_pdf/<int:quote_id>')
@login_required
def export_quote_pdf(quote_id):
    """Export a quote as PDF"""
    with get_db() as conn:
        quote = conn.execute("""
            SELECT q.id, q.description, q.amount, q.status, q.created_at,
                   cu.name, cu.phone, cu.email
            FROM quotes q JOIN customers cu ON q.customer_id=cu.id
            WHERE q.id=?
        """, (quote_id,)).fetchone()
    if not quote:
        flash("Devis introuvable", "error")
        return redirect("/quotes")

    settings = get_all_settings()
    shop_name = settings.get('shop_name', 'AMILCAR Auto Care')
    shop_phone = settings.get('shop_phone', '')
    shop_address = settings.get('shop_address', '')
    tax_rate = float(settings.get('tax_rate', '0') or '0')
    today = date.today()

    amount = float(quote['amount'])
    tax_amount = round(amount * tax_rate / 100, 2)
    total_ttc = round(amount + tax_amount, 2)

    html = f"""<!DOCTYPE html>
<html><head><meta charset="utf-8">
<style>
  body {{ font-family: Arial, sans-serif; font-size: 11px; color: #1a1a2e; margin: 24px; }}
  .header {{ background: #1a1a2e; color: white; padding: 18px; display: flex; justify-content: space-between; }}
  .header h1 {{ color: #e8c547; margin: 0; font-size: 18px; letter-spacing: 3px; }}
  .header p {{ color: #aaa; margin: 4px 0 0; font-size: 9px; }}
  .title {{ text-align: center; color: #B8962E; font-size: 16px; letter-spacing: 6px; margin: 24px 0; font-weight: 700; }}
  .info {{ display: table; width: 100%; margin: 14px 0; }}
  .info-col {{ display: table-cell; width: 50%; padding: 12px; vertical-align: top; border: 1px solid #e8e5dd; background: #fafaf8; }}
  .info-col h3 {{ color: #B8962E; font-size: 9px; letter-spacing: 2px; margin: 0 0 8px; }}
  .info-col p {{ margin: 3px 0; font-size: 11px; }}
  .desc {{ background: #f8f9fa; border: 1px solid #e0e0e0; padding: 16px; margin: 18px 0; border-radius: 4px; font-size: 12px; line-height: 1.6; }}
  .totals {{ width: 50%; margin-left: auto; margin-top: 18px; }}
  .totals tr td {{ padding: 8px 12px; font-size: 11px; }}
  .totals .total-row td {{ background: #1a1a2e; color: #e8c547; font-size: 15px; font-weight: 800; }}
  .validity {{ text-align: center; color: #888; font-size: 10px; margin: 24px 0; padding: 10px; border: 1px dashed #ddd; }}
  .footer {{ text-align: center; color: #aaa; font-size: 8px; margin-top: 30px; border-top: 1px solid #ddd; padding-top: 8px; }}
</style></head><body>
<div class="header">
  <div><h1>{shop_name}</h1><p>{shop_address}</p><p>{shop_phone}</p></div>
  <div style="text-align:right"><p>Date: {today.strftime('%d/%m/%Y')}</p><p>Devis N° {quote['id']}</p></div>
</div>
<div class="title">DEVIS</div>
<div class="info">
  <div class="info-col"><h3>CLIENT</h3><p><b>{quote['name']}</b></p><p>{quote['phone'] or ''}</p><p>{quote['email'] or ''}</p></div>
  <div class="info-col"><h3>DÉTAILS</h3><p>Devis N°: <b>#{quote['id']}</b></p><p>Date: {quote['created_at'][:10] if quote['created_at'] else today}</p><p>Statut: <b>{quote['status']}</b></p></div>
</div>
<div class="desc"><b>Description:</b><br>{quote['description'] or '—'}</div>
<table class="totals">
  <tr><td>Montant HT</td><td style="text-align:right">{round(amount,2)} DT</td></tr>
  <tr><td>TVA ({tax_rate}%)</td><td style="text-align:right">{tax_amount} DT</td></tr>
  <tr class="total-row"><td>TOTAL TTC</td><td style="text-align:right">{total_ttc} DT</td></tr>
</table>
<div class="validity">Ce devis est valable 30 jours à compter de sa date d'émission.</div>
<div class="footer">{shop_name} — {shop_address} — {shop_phone}<br>Document confidentiel — Généré le {today.strftime('%d/%m/%Y à %H:%M')}</div>
</body></html>"""

    buf = _render_pdf(html)
    log_activity('Export', f'Devis #{quote_id} → PDF')
    return send_file(buf, as_attachment=True,
                     download_name=f"devis_{quote_id}_{today}.pdf",
                     mimetype="application/pdf")


# ─── P&L Report with Year-over-Year Comparison ──────────────────────────────
@exports_bp.route("/export/financial_excel")
@login_required
def export_financial_excel():
    """Export full financial report as multi-sheet Excel."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    year = request.args.get('year', date.today().year, type=int)
    wb = Workbook()
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1B2A4A", end_color="1B2A4A", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'))
    month_names = ["JAN","FEB","MAR","APR","MAY","JUN","JUL","AUG","SEP","OCT","NOV","DEC"]

    with get_db() as conn:
        # Sheet 1: P&L Monthly
        ws = wb.active
        ws.title = "P&L Mensuel"
        headers = ["Mois", "Revenu", "Dépenses", "Profit", "Marge %",
                    f"Revenu {year-1}", f"Dépenses {year-1}", f"Profit {year-1}", "Évolution %"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border

        for m in range(1, 13):
            ms = f"{year}-{m:02d}-01"
            me = f"{year}-{m+1:02d}-01" if m < 12 else f"{year+1}-01-01"
            pms = f"{year-1}-{m:02d}-01"
            pme = f"{year-1}-{m+1:02d}-01" if m < 12 else f"{year}-01-01"
            rev = conn.execute("SELECT COALESCE(SUM(i.amount),0) FROM invoices i JOIN appointments a ON i.appointment_id=a.id WHERE a.date>=? AND a.date<? AND i.status='paid'", (ms, me)).fetchone()[0]
            exp = conn.execute("SELECT COALESCE(SUM(amount),0) FROM expenses WHERE date>=? AND date<?", (ms, me)).fetchone()[0]
            prev_rev = conn.execute("SELECT COALESCE(SUM(i.amount),0) FROM invoices i JOIN appointments a ON i.appointment_id=a.id WHERE a.date>=? AND a.date<? AND i.status='paid'", (pms, pme)).fetchone()[0]
            prev_exp = conn.execute("SELECT COALESCE(SUM(amount),0) FROM expenses WHERE date>=? AND date<?", (pms, pme)).fetchone()[0]
            profit = rev - exp
            margin = round(profit / rev * 100, 1) if rev > 0 else 0
            yoy = round((rev - prev_rev) / prev_rev * 100, 1) if prev_rev > 0 else 0
            row = [month_names[m-1], rev, exp, profit, margin, prev_rev, prev_exp, prev_rev - prev_exp, yoy]
            for col, val in enumerate(row, 1):
                cell = ws.cell(row=m+1, column=col, value=val)
                cell.border = thin_border
                if col >= 2:
                    cell.number_format = '#,##0.000' if isinstance(val, float) and col in (5, 9) else '#,##0'

        for col in range(1, 10):
            ws.column_dimensions[chr(64+col)].width = 16

        # Sheet 2: Expense Categories
        ws2 = wb.create_sheet("Dépenses par Catégorie")
        ws2_headers = ["Catégorie", "Montant Total", "% du Total"]
        for col, h in enumerate(ws2_headers, 1):
            cell = ws2.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
        cats = conn.execute(
            "SELECT COALESCE(category,'Autre'), COALESCE(SUM(amount),0) "
            "FROM expenses WHERE strftime('%%Y',date)=? GROUP BY category ORDER BY SUM(amount) DESC",
            (str(year),)).fetchall()
        total_exp_all = sum(c[1] for c in cats) if cats else 1
        for r, cat in enumerate(cats, 2):
            ws2.cell(row=r, column=1, value=cat[0]).border = thin_border
            ws2.cell(row=r, column=2, value=cat[1]).border = thin_border
            ws2.cell(row=r, column=3, value=round(cat[1]/total_exp_all*100, 1)).border = thin_border
        for col in range(1, 4):
            ws2.column_dimensions[chr(64+col)].width = 22

        # Sheet 3: Top Services
        ws3 = wb.create_sheet("Top Services")
        ws3_headers = ["Service", "Nombre", "Revenu Total"]
        for col, h in enumerate(ws3_headers, 1):
            cell = ws3.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
        svcs = conn.execute(
            "SELECT a.service, COUNT(*), COALESCE(SUM(i.amount),0) "
            "FROM appointments a LEFT JOIN invoices i ON i.appointment_id=a.id AND i.status='paid' "
            "GROUP BY a.service ORDER BY SUM(i.amount) DESC LIMIT 20").fetchall()
        for r, s in enumerate(svcs, 2):
            for col, val in enumerate(s, 1):
                ws3.cell(row=r, column=col, value=val).border = thin_border
        for col in range(1, 4):
            ws3.column_dimensions[chr(64+col)].width = 30

        # Sheet 4: Top Customers
        ws4 = wb.create_sheet("Top Clients")
        ws4_headers = ["Client", "Total Payé", "Visites"]
        for col, h in enumerate(ws4_headers, 1):
            cell = ws4.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
        custs = conn.execute(
            "SELECT cu.name, COALESCE(SUM(i.amount),0), COUNT(DISTINCT a.id) "
            "FROM customers cu JOIN cars ca ON ca.customer_id=cu.id "
            "JOIN appointments a ON a.car_id=ca.id "
            "LEFT JOIN invoices i ON i.appointment_id=a.id AND i.status='paid' "
            "GROUP BY cu.id ORDER BY SUM(i.amount) DESC LIMIT 20").fetchall()
        for r, c in enumerate(custs, 2):
            for col, val in enumerate(c, 1):
                ws4.cell(row=r, column=col, value=val).border = thin_border
        for col in range(1, 4):
            ws4.column_dimensions[chr(64+col)].width = 25

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    log_activity('Export', f'Financial Excel {year}')
    return send_file(buf, as_attachment=True,
                     download_name=f"rapport_financier_{year}.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


# ─── WhatsApp Auto Reminders ────────────────────────────────────────────────
