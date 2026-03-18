"""
AMILCAR — Customer Management & CRM
Blueprint: customers_bp
Routes: 42
"""
from flask import Blueprint, render_template, request, redirect, url_for, flash, make_response, jsonify, session, send_file
from helpers import login_required, admin_required, client_required, get_db, get_services, get_setting, get_all_settings
from helpers import allowed_file, safe_page, log_activity, build_wa_url, STATUS_MESSAGES, UPLOAD_FOLDER, MAX_FILE_SIZE, MAX_FILES, PER_PAGE
from database.db import get_db
from werkzeug.utils import secure_filename
from werkzeug.security import generate_password_hash, check_password_hash
from datetime import datetime, date, timedelta
import os, re, uuid, io
import time as time_module
import sqlite3

customers_bp = Blueprint("customers_bp", __name__)

LOYALTY_THRESHOLD = 5
TIER_THRESHOLDS = {'BRONZE': 0, 'ARGENT': 500, 'OR': 1000, 'PLATINE': 2000}
RFM_SEGMENTS = {
    (5,5): 'champion', (5,4): 'champion', (4,5): 'champion',
    (5,3): 'loyal', (4,4): 'loyal', (3,5): 'loyal',
    (5,2): 'potential_loyalist', (4,3): 'potential_loyalist', (3,4): 'potential_loyalist',
    (5,1): 'new_customer', (4,1): 'new_customer', (4,2): 'new_customer',
    (3,3): 'need_attention', (3,2): 'need_attention', (2,3): 'need_attention',
    (3,1): 'about_to_sleep', (2,2): 'about_to_sleep',
    (2,1): 'at_risk', (1,3): 'at_risk', (1,4): 'at_risk', (1,5): 'at_risk',
    (1,2): 'hibernating', (1,1): 'lost', (2,4): 'at_risk', (2,5): 'at_risk',
}
RFM_LABELS = {
    'champion': ('Champion', '#34d399'), 'loyal': ('Fidèle', '#D4AF37'),
    'potential_loyalist': ('Potentiel Fidèle', '#1B6B93'), 'new_customer': ('Nouveau', '#60a5fa'),
    'need_attention': ('Attention requise', '#f59e0b'), 'about_to_sleep': ('En veille', '#f97316'),
    'at_risk': ('À risque', '#ef4444'), 'hibernating': ('Hibernant', '#6b7280'),
    'lost': ('Perdu', '#374151'), 'new': ('Non classé', '#9ca3af'),
}


@customers_bp.route('/customers')
@login_required
def customers():
    search = request.args.get('q', '').strip()
    page = safe_page(request.args.get('page', 1, type=int))
    with get_db() as conn:
        if search:
            total = conn.execute(
                "SELECT COUNT(*) FROM customers WHERE COALESCE(is_deleted,0)=0 AND (name LIKE ? OR phone LIKE ?)",
                (f'%{search}%', f'%{search}%')
            ).fetchone()[0]
            all_customers = conn.execute(
                "SELECT * FROM customers WHERE COALESCE(is_deleted,0)=0 AND (name LIKE ? OR phone LIKE ?) LIMIT ? OFFSET ?",
                (f'%{search}%', f'%{search}%', PER_PAGE, (page - 1) * PER_PAGE)
            ).fetchall()
        else:
            total = conn.execute("SELECT COUNT(*) FROM customers WHERE COALESCE(is_deleted,0)=0").fetchone()[0]
            all_customers = conn.execute(
                "SELECT * FROM customers WHERE COALESCE(is_deleted,0)=0 LIMIT ? OFFSET ?",
                (PER_PAGE, (page - 1) * PER_PAGE)
            ).fetchall()
    total_pages = (total + PER_PAGE - 1) // PER_PAGE
    return render_template('customers.html', customers=all_customers, search=search,
                           page=page, total_pages=total_pages)



@customers_bp.route("/customer/<int:customer_id>")
@login_required
def customer_detail(customer_id):
    with get_db() as conn:
        customer = conn.execute("SELECT * FROM customers WHERE id = ?", (customer_id,)).fetchone()
        if not customer:
            flash('Client introuvable', 'error')
            return redirect("/customers")
        cars = conn.execute("SELECT * FROM cars WHERE customer_id = ? AND COALESCE(is_deleted,0)=0", (customer_id,)).fetchall()
        appointments = conn.execute("SELECT a.* FROM appointments a JOIN cars c ON a.car_id = c.id WHERE c.customer_id = ? AND COALESCE(a.is_deleted,0)=0 ORDER BY a.id DESC", (customer_id,)).fetchall()
        # CLV calculation
        total_spent = conn.execute(
            "SELECT COALESCE(SUM(i.amount),0) FROM invoices i JOIN appointments a ON i.appointment_id = a.id "
            "JOIN cars ca ON a.car_id = ca.id WHERE ca.customer_id = ? AND i.status = 'paid'",
            (customer_id,)).fetchone()[0]
        visit_count = conn.execute(
            "SELECT COUNT(*) FROM appointments a JOIN cars ca ON a.car_id = ca.id "
            "WHERE ca.customer_id = ? AND a.status = 'completed'", (customer_id,)).fetchone()[0]
        first_visit = conn.execute(
            "SELECT MIN(a.date) FROM appointments a JOIN cars ca ON a.car_id = ca.id "
            "WHERE ca.customer_id = ?", (customer_id,)).fetchone()[0]
        # Average rating
        avg_rating = conn.execute(
            "SELECT AVG(rating), COUNT(*) FROM ratings WHERE customer_id = ?",
            (customer_id,)).fetchone()
        # CLV tier
        if total_spent >= 1000:
            tier = 'OR'
        elif total_spent >= 500:
            tier = 'ARGENT'
        elif total_spent >= 200:
            tier = 'BRONZE'
        else:
            tier = '—'
        clv = {
            'total_spent': total_spent, 'visits': visit_count,
            'first_visit': first_visit or '—',
            'avg_rating': round(avg_rating[0], 1) if avg_rating[0] else 0,
            'rating_count': avg_rating[1],
            'tier': tier
        }
    return render_template("customer_detail.html", customer=customer, cars=cars,
                           appointments=appointments, clv=clv)



@customers_bp.route("/export/customers")
@login_required
def export_customers_csv():
    import csv
    with get_db() as conn:
        customers = conn.execute("SELECT id, name, phone, notes FROM customers ORDER BY id").fetchall()
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["ID", "Nom", "Téléphone", "Notes"])
    for c in customers:
        writer.writerow(c)
    response = make_response(output.getvalue())
    response.headers['Content-Type'] = 'text/csv; charset=utf-8'
    response.headers['Content-Disposition'] = 'attachment; filename=customers.csv'
    return response



@customers_bp.route("/loyalty")
@login_required
def loyalty_page():
    with get_db() as conn:
        loyalty_data = conn.execute(
            "SELECT l.id, cu.id, cu.name, cu.phone, l.service_type, l.wash_count, l.free_washes_used "
            "FROM loyalty l JOIN customers cu ON l.customer_id = cu.id "
            "ORDER BY l.wash_count DESC"
        ).fetchall()
        all_customers = conn.execute("SELECT id, name, phone FROM customers ORDER BY name").fetchall()
    entries = []
    for row in loyalty_data:
        washes = row[5]
        free_used = row[6]
        free_earned = washes // LOYALTY_THRESHOLD
        free_available = free_earned - free_used
        entries.append({
            'id': row[0], 'customer_id': row[1], 'name': row[2], 'phone': row[3],
            'service_type': row[4], 'wash_count': washes,
            'free_earned': free_earned, 'free_used': free_used,
            'free_available': max(0, free_available),
            'progress': washes % LOYALTY_THRESHOLD
        })
    return render_template("loyalty.html", entries=entries, threshold=LOYALTY_THRESHOLD, customers=all_customers)



@customers_bp.route("/loyalty/add_wash", methods=["POST"])
@login_required
def loyalty_add_wash():
    customer_id = request.form.get("customer_id", "")
    service_type = request.form.get("service_type", "Lavage Normal")
    if not customer_id:
        flash("Sélectionnez un client", "error")
        return redirect("/loyalty")
    with get_db() as conn:
        existing = conn.execute(
            "SELECT id, wash_count FROM loyalty WHERE customer_id = ? AND service_type = ?",
            (customer_id, service_type)).fetchone()
        if existing:
            new_count = existing[1] + 1
            conn.execute("UPDATE loyalty SET wash_count = ? WHERE id = ?", (new_count, existing[0]))
        else:
            conn.execute("INSERT INTO loyalty (customer_id, service_type, wash_count) VALUES (?, ?, 1)",
                (customer_id, service_type))
            new_count = 1
        conn.commit()
    if new_count % LOYALTY_THRESHOLD == 0:
        flash(f"🎉 Le client a gagné un lavage GRATUIT ! (lavage #{new_count})", "success")
    else:
        remaining = LOYALTY_THRESHOLD - (new_count % LOYALTY_THRESHOLD)
        flash(f"Lavage #{new_count} enregistré. Encore {remaining} pour un lavage gratuit !", "success")
    log_activity('Loyalty Wash', f'Customer #{customer_id} — {service_type} (#{new_count})')
    return redirect("/loyalty")



@customers_bp.route("/loyalty/use_free", methods=["POST"])
@login_required
def loyalty_use_free():
    loyalty_id = request.form.get("loyalty_id", "")
    with get_db() as conn:
        row = conn.execute("SELECT wash_count, free_washes_used FROM loyalty WHERE id = ?", (loyalty_id,)).fetchone()
        if row:
            free_earned = row[0] // LOYALTY_THRESHOLD
            if row[1] < free_earned:
                conn.execute("UPDATE loyalty SET free_washes_used = free_washes_used + 1 WHERE id = ?", (loyalty_id,))
                conn.commit()
                flash("Lavage gratuit utilisé avec succès !", "success")
            else:
                flash("Aucun lavage gratuit disponible", "error")
    return redirect("/loyalty")



# ─── Customer Report Export ───
@customers_bp.route("/customer_report/<int:customer_id>")
@login_required
def customer_report(customer_id):
    import csv
    with get_db() as conn:
        customer = conn.execute("SELECT * FROM customers WHERE id = ?", (customer_id,)).fetchone()
        if not customer:
            flash('Client introuvable', 'error')
            return redirect("/customers")
        cars = conn.execute("SELECT * FROM cars WHERE customer_id = ?", (customer_id,)).fetchall()
        appointments = conn.execute(
            "SELECT a.id, a.date, a.service, a.status, ca.brand, ca.model "
            "FROM appointments a JOIN cars ca ON a.car_id = ca.id "
            "WHERE ca.customer_id = ? ORDER BY a.date DESC", (customer_id,)).fetchall()
        invoices = conn.execute(
            "SELECT i.id, a.date, a.service, i.amount, i.status, i.payment_method "
            "FROM invoices i JOIN appointments a ON i.appointment_id = a.id "
            "JOIN cars ca ON a.car_id = ca.id WHERE ca.customer_id = ? ORDER BY a.date DESC", (customer_id,)).fetchall()
        total_spent = conn.execute(
            "SELECT COALESCE(SUM(i.amount),0) FROM invoices i JOIN appointments a ON i.appointment_id = a.id "
            "JOIN cars ca ON a.car_id = ca.id WHERE ca.customer_id = ? AND i.status = 'paid'", (customer_id,)).fetchone()[0]
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow([f"Rapport Client — {customer[1]}"])
    writer.writerow([f"Téléphone: {customer[2]}", f"Email: {customer[4] if len(customer) > 4 else ''}"])
    writer.writerow([f"Total Dépensé: {total_spent} DT"])
    writer.writerow([])
    writer.writerow(["=== VOITURES ==="])
    writer.writerow(["Marque", "Modèle", "Plaque", "Année", "Couleur"])
    for c in cars:
        writer.writerow([c[2], c[3], c[4], c[5] if len(c) > 5 else '', c[6] if len(c) > 6 else ''])
    writer.writerow([])
    writer.writerow(["=== RENDEZ-VOUS ==="])
    writer.writerow(["ID", "Date", "Service", "Statut", "Voiture"])
    for a in appointments:
        writer.writerow([a[0], a[1], a[2], a[3], f"{a[4]} {a[5]}"])
    writer.writerow([])
    writer.writerow(["=== FACTURES ==="])
    writer.writerow(["ID", "Date", "Service", "Montant (DT)", "Statut", "Paiement"])
    for inv in invoices:
        writer.writerow(inv)
    response = make_response(output.getvalue())
    response.headers['Content-Type'] = 'text/csv; charset=utf-8'
    safe_name = customer[1].replace(' ', '_')
    response.headers['Content-Disposition'] = f'attachment; filename=customer_report_{safe_name}.csv'
    return response



# ─── Excel Import/Export ───
@customers_bp.route("/export/customers_csv_full")
@login_required
def export_customers_csv_full():
    import csv
    with get_db() as conn:
        customers = conn.execute(
            "SELECT c.id, c.name, c.phone, COALESCE(c.email,''), COALESCE(c.notes,''), "
            "(SELECT COUNT(*) FROM cars WHERE customer_id = c.id), "
            "(SELECT COALESCE(SUM(i.amount),0) FROM invoices i JOIN appointments a ON i.appointment_id = a.id "
            "JOIN cars ca ON a.car_id = ca.id WHERE ca.customer_id = c.id AND i.status = 'paid') "
            "FROM customers c ORDER BY c.name").fetchall()
    output = io.StringIO()
    writer = csv.writer(output, delimiter=';')
    writer.writerow(["ID", "Nom", "Téléphone", "Email", "Notes", "Nb Voitures", "Total Payé (DT)"])
    for c in customers:
        writer.writerow(c)
    response = make_response('\ufeff' + output.getvalue())
    response.headers['Content-Type'] = 'text/csv; charset=utf-8-sig'
    response.headers['Content-Disposition'] = 'attachment; filename=clients_amilcar.csv'
    return response



@customers_bp.route("/import/customers", methods=["POST"])
@admin_required
def import_customers():
    import csv
    file = request.files.get("file")
    if not file or not file.filename:
        flash("Sélectionnez un fichier", "error")
        return redirect("/customers")
    fname = file.filename.lower()
    if not fname.endswith(('.csv', '.txt', '.xlsx')):
        flash("Format non supporté. Utilisez CSV ou Excel (.xlsx)", "error")
        return redirect("/customers")
    try:
        rows = []
        if fname.endswith('.xlsx'):
            from openpyxl import load_workbook
            wb = load_workbook(filename=io.BytesIO(file.read()), read_only=True, data_only=True)
            ws = wb.active
            for data_row in ws.iter_rows(min_row=2, values_only=True):
                rows.append([str(c).strip() if c is not None else '' for c in data_row])
            wb.close()
        else:
            content = file.read().decode('utf-8-sig')
            reader = csv.reader(io.StringIO(content), delimiter=';')
            header = next(reader, None)
            if not header:
                flash("Fichier vide", "error")
                return redirect("/customers")
            rows = list(reader)
        imported = 0
        skipped = 0
        with get_db() as conn:
            for row in reader:
                if len(row) < 2:
                    skipped += 1
                    continue
                name = row[0].strip() if not row[0].strip().isdigit() else (row[1].strip() if len(row) > 1 else '')
                phone = row[1].strip() if not row[0].strip().isdigit() else (row[2].strip() if len(row) > 2 else '')
                # If first column is ID (number), shift
                if row[0].strip().isdigit() and len(row) >= 3:
                    name = row[1].strip()
                    phone = row[2].strip()
                if not name or not phone:
                    skipped += 1
                    continue
                existing = conn.execute("SELECT id FROM customers WHERE phone = ?", (phone,)).fetchone()
                if existing:
                    skipped += 1
                    continue
                email = ''
                notes = ''
                if len(row) > 3:
                    email = row[3].strip()
                if len(row) > 4:
                    notes = row[4].strip()
                conn.execute("INSERT INTO customers (name, phone, email, notes) VALUES (?,?,?,?)",
                    (name, phone, email, notes))
                imported += 1
            conn.commit()
        log_activity('Import', f'{imported} clients importés, {skipped} ignorés')
        flash(f"{imported} clients importés, {skipped} ignorés (doublons/invalides)", "success")
    except Exception as e:
        flash(f"Erreur d'import : {str(e)}", "error")
    return redirect("/customers")



@customers_bp.route("/rewards")
@login_required
def rewards_page():
    with get_db() as conn:
        rewards = conn.execute(
            "SELECT rp.id, cu.id, cu.name, cu.phone, rp.points, rp.total_earned, rp.total_spent, rp.tier "
            "FROM reward_points rp JOIN customers cu ON rp.customer_id = cu.id "
            "ORDER BY rp.points DESC").fetchall()
        customers = conn.execute("SELECT id, name, phone FROM customers ORDER BY name").fetchall()
    return render_template("rewards.html", rewards=rewards, customers=customers,
                           tiers=TIER_THRESHOLDS)



@customers_bp.route("/rewards/add_points", methods=["POST"])
@login_required
def add_reward_points():
    customer_id = request.form.get("customer_id", "")
    points = request.form.get("points", "0")
    description = request.form.get("description", "").strip()
    if not customer_id:
        flash("Sélectionnez un client", "error")
        return redirect("/rewards")
    try:
        pts = int(points)
        if pts <= 0:
            raise ValueError
    except ValueError:
        flash("Nombre de points invalide", "error")
        return redirect("/rewards")
    with get_db() as conn:
        existing = conn.execute("SELECT id, points, total_earned FROM reward_points WHERE customer_id = ?",
            (customer_id,)).fetchone()
        if existing:
            new_points = existing[1] + pts
            new_total = existing[2] + pts
            tier = _calculate_tier(new_total)
            conn.execute("UPDATE reward_points SET points = ?, total_earned = ?, tier = ? WHERE id = ?",
                (new_points, new_total, tier, existing[0]))
        else:
            tier = _calculate_tier(pts)
            conn.execute("INSERT INTO reward_points (customer_id, points, total_earned, tier) VALUES (?,?,?,?)",
                (customer_id, pts, pts, tier))
        conn.execute("INSERT INTO reward_history (customer_id, points, type, description) VALUES (?,?,?,?)",
            (customer_id, pts, 'earn', description or f'+{pts} points'))
        conn.commit()
    flash(f"{pts} points ajoutés", "success")
    return redirect("/rewards")



@customers_bp.route("/rewards/redeem", methods=["POST"])
@login_required
def redeem_reward_points():
    customer_id = request.form.get("customer_id", "")
    points = request.form.get("points", "0")
    reward_desc = request.form.get("reward", "").strip()
    if not customer_id:
        flash("Client requis", "error")
        return redirect("/rewards")
    try:
        pts = int(points)
        if pts <= 0:
            raise ValueError
    except ValueError:
        flash("Nombre de points invalide", "error")
        return redirect("/rewards")
    with get_db() as conn:
        existing = conn.execute("SELECT id, points, total_spent FROM reward_points WHERE customer_id = ?",
            (customer_id,)).fetchone()
        if not existing or existing[1] < pts:
            flash("Points insuffisants", "error")
            return redirect("/rewards")
        conn.execute("UPDATE reward_points SET points = points - ?, total_spent = total_spent + ? WHERE id = ?",
            (pts, pts, existing[0]))
        conn.execute("INSERT INTO reward_history (customer_id, points, type, description) VALUES (?,?,?,?)",
            (customer_id, -pts, 'redeem', reward_desc or f'Échange {pts} points'))
        conn.commit()
    flash(f"{pts} points échangés", "success")
    return redirect("/rewards")



# ─── Phase 6 Feature 7: Customer Analytics ───
@customers_bp.route("/customer_analytics/<int:customer_id>")
@login_required
def customer_analytics(customer_id):
    with get_db() as conn:
        customer = conn.execute("SELECT * FROM customers WHERE id = ?", (customer_id,)).fetchone()
        if not customer:
            flash("Client introuvable", "error")
            return redirect("/customers")
        cars = conn.execute("SELECT * FROM cars WHERE customer_id = ?", (customer_id,)).fetchall()
        car_ids = [c[0] for c in cars]
        if car_ids:
            placeholders = ','.join(['?' for _ in car_ids])
            appointments = conn.execute(
                f"SELECT a.id, a.date, a.service, a.status, ca.brand, ca.model, ca.plate "
                f"FROM appointments a JOIN cars ca ON a.car_id=ca.id WHERE ca.customer_id=? ORDER BY a.date DESC", (customer_id,)).fetchall()
            total_spent = conn.execute(
                f"SELECT COALESCE(SUM(i.amount),0) FROM invoices i JOIN appointments a ON i.appointment_id=a.id "
                f"JOIN cars ca ON a.car_id=ca.id WHERE ca.customer_id=? AND i.status='paid'", (customer_id,)).fetchone()[0]
            visit_count = conn.execute(
                f"SELECT COUNT(*) FROM appointments a JOIN cars ca ON a.car_id=ca.id WHERE ca.customer_id=?", (customer_id,)).fetchone()[0]
            services_used = conn.execute(
                f"SELECT a.service, COUNT(*) as cnt FROM appointments a JOIN cars ca ON a.car_id=ca.id "
                f"WHERE ca.customer_id=? GROUP BY a.service ORDER BY cnt DESC", (customer_id,)).fetchall()
            monthly_spending = conn.execute(
                f"SELECT strftime('%Y-%m', a.date) as m, COALESCE(SUM(i.amount),0) "
                f"FROM appointments a JOIN cars ca ON a.car_id=ca.id LEFT JOIN invoices i ON i.appointment_id=a.id AND i.status='paid' "
                f"WHERE ca.customer_id=? GROUP BY m ORDER BY m DESC LIMIT 12", (customer_id,)).fetchall()
            first_visit = conn.execute(
                f"SELECT MIN(a.date) FROM appointments a JOIN cars ca ON a.car_id=ca.id WHERE ca.customer_id=?", (customer_id,)).fetchone()[0]
            last_visit = conn.execute(
                f"SELECT MAX(a.date) FROM appointments a JOIN cars ca ON a.car_id=ca.id WHERE ca.customer_id=?", (customer_id,)).fetchone()[0]
        else:
            appointments, total_spent, visit_count, services_used, monthly_spending = [], 0, 0, [], []
            first_visit, last_visit = None, None
        # Rewards info
        rewards = conn.execute("SELECT points, total_earned, tier FROM reward_points WHERE customer_id=?", (customer_id,)).fetchone()
        # Ratings
        avg_rating = conn.execute(
            "SELECT AVG(r.rating) FROM ratings r WHERE r.customer_id=?", (customer_id,)).fetchone()[0]
    # Predict next visit
    from datetime import date, timedelta
    predicted_next = None
    if visit_count >= 2 and last_visit and first_visit:
        try:
            d_first = date.fromisoformat(first_visit)
            d_last = date.fromisoformat(last_visit)
            avg_gap = (d_last - d_first).days / max(visit_count - 1, 1)
            predicted_next = (d_last + timedelta(days=int(avg_gap))).isoformat()
        except (ValueError, TypeError):
            pass
    return render_template("customer_analytics.html",
        customer=customer, cars=cars, appointments=appointments[:20],
        total_spent=total_spent, visit_count=visit_count,
        services_used=services_used, monthly_spending=list(reversed(monthly_spending)),
        first_visit=first_visit, last_visit=last_visit,
        predicted_next=predicted_next, rewards=rewards,
        avg_rating=round(avg_rating, 1) if avg_rating else None)



# ─── Phase 7 Feature 7: Customer Satisfaction Survey ───
@customers_bp.route("/surveys")
@login_required
def surveys_list():
    with get_db() as conn:
        surveys = conn.execute("""
            SELECT s.*, c.name, a.service, a.date
            FROM surveys s
            JOIN customers c ON s.customer_id = c.id
            JOIN appointments a ON s.appointment_id = a.id
            ORDER BY s.created_at DESC
        """).fetchall()
        # Stats
        submitted = [s for s in surveys if s[11]]
        avg_quality = sum(s[4] for s in submitted) / len(submitted) if submitted else 0
        avg_speed = sum(s[5] for s in submitted) / len(submitted) if submitted else 0
        avg_reception = sum(s[6] for s in submitted) / len(submitted) if submitted else 0
        avg_cleanliness = sum(s[7] for s in submitted) / len(submitted) if submitted else 0
        avg_value = sum(s[8] for s in submitted) / len(submitted) if submitted else 0
        avg_overall = (avg_quality + avg_speed + avg_reception + avg_cleanliness + avg_value) / 5 if submitted else 0
    return render_template("surveys.html", surveys=surveys,
        avg_quality=round(avg_quality,1), avg_speed=round(avg_speed,1),
        avg_reception=round(avg_reception,1), avg_cleanliness=round(avg_cleanliness,1),
        avg_value=round(avg_value,1), avg_overall=round(avg_overall,1),
        total_submitted=len(submitted), total_pending=len(surveys)-len(submitted))



@customers_bp.route("/survey/create/<int:appointment_id>", methods=["POST"])
@login_required
def create_survey(appointment_id):
    with get_db() as conn:
        appt = conn.execute("SELECT a.*, cr.customer_id FROM appointments a JOIN cars cr ON a.car_id=cr.id WHERE a.id=?", (appointment_id,)).fetchone()
        if not appt:
            flash("Rendez-vous introuvable", "danger")
            return redirect("/appointments")
        existing = conn.execute("SELECT id FROM surveys WHERE appointment_id=?", (appointment_id,)).fetchone()
        if existing:
            flash("Un questionnaire existe déjà pour ce RDV", "warning")
            return redirect("/surveys")
        token = uuid.uuid4().hex[:12]
        customer_id = appt[-1]
        conn.execute("INSERT INTO surveys (appointment_id, customer_id, token) VALUES (?,?,?)",
                    (appointment_id, customer_id, token))
        conn.commit()
    flash(f"Questionnaire créé ! Lien: /survey/{token}", "success")
    return redirect("/surveys")



@customers_bp.route("/survey/<token>", methods=["GET", "POST"])
def fill_survey(token):
    with get_db() as conn:
        survey = conn.execute("SELECT s.*, c.name, a.service FROM surveys s JOIN customers c ON s.customer_id=c.id JOIN appointments a ON s.appointment_id=a.id WHERE s.token=?", (token,)).fetchone()
        if not survey:
            return "Questionnaire introuvable", 404
        if survey[11]:  # already submitted
            return render_template("survey_thanks.html", survey=survey)
        if request.method == "POST":
            from datetime import datetime
            conn.execute("""UPDATE surveys SET
                q_quality=?, q_speed=?, q_reception=?, q_cleanliness=?, q_value=?,
                comment=?, submitted=1, submitted_at=? WHERE token=?""",
                (int(request.form.get('q_quality', 3)),
                 int(request.form.get('q_speed', 3)),
                 int(request.form.get('q_reception', 3)),
                 int(request.form.get('q_cleanliness', 3)),
                 int(request.form.get('q_value', 3)),
                 request.form.get('comment', ''),
                 datetime.now().isoformat(), token))
            conn.commit()
            return render_template("survey_thanks.html", survey=survey)
    return render_template("survey_form.html", survey=survey)



# ─── 3. CRM Follow-up System ───
@customers_bp.route("/crm_followups")
@login_required
def crm_followups():
    with get_db() as conn:
        followups = conn.execute("""SELECT f.*, c.name, c.phone FROM crm_followups f
            JOIN customers c ON f.customer_id=c.id ORDER BY f.scheduled_date""").fetchall()
        # Clients absents > 60 jours
        from datetime import date, timedelta
        cutoff = (date.today() - timedelta(days=60)).isoformat()
        absent = conn.execute("""SELECT c.id, c.name, c.phone, MAX(a.date) as last_visit
            FROM customers c LEFT JOIN cars ca ON ca.customer_id=c.id
            LEFT JOIN appointments a ON a.car_id=ca.id
            GROUP BY c.id HAVING last_visit < ? OR last_visit IS NULL
            ORDER BY last_visit""", (cutoff,)).fetchall()
    return render_template("crm_followups.html", followups=followups, absent=absent)



@customers_bp.route("/crm_followup/add", methods=["POST"])
@login_required
def crm_followup_add():
    cid = request.form.get("customer_id", type=int)
    ftype = request.form.get("type", "absence")
    scheduled = request.form.get("scheduled_date", "")
    reason = request.form.get("reason", "")
    notes = request.form.get("notes", "")
    if cid and scheduled:
        with get_db() as conn:
            conn.execute("INSERT INTO crm_followups (customer_id, type, scheduled_date, reason, notes) VALUES (?,?,?,?,?)",
                        (cid, ftype, scheduled, reason, notes))
            conn.commit()
        flash("Suivi CRM ajouté !", "success")
    return redirect("/crm_followups")



@customers_bp.route("/crm_followup/complete/<int:fid>", methods=["POST"])
@login_required
def crm_followup_complete(fid):
    from datetime import date
    with get_db() as conn:
        conn.execute("UPDATE crm_followups SET status='completed', completed_at=? WHERE id=?",
                    (date.today().isoformat(), fid))
        conn.commit()
    flash("Suivi marqué comme complété", "success")
    return redirect("/crm_followups")



@customers_bp.route("/crm_followup/auto_generate", methods=["POST"])
@login_required
def crm_followup_auto():
    from datetime import date, timedelta
    cutoff = (date.today() - timedelta(days=60)).isoformat()
    scheduled = (date.today() + timedelta(days=1)).isoformat()
    with get_db() as conn:
        absent = conn.execute("""SELECT c.id FROM customers c LEFT JOIN cars ca ON ca.customer_id=c.id
            LEFT JOIN appointments a ON a.car_id=ca.id GROUP BY c.id
            HAVING MAX(a.date) < ? OR MAX(a.date) IS NULL""", (cutoff,)).fetchall()
        existing = set(r[0] for r in conn.execute(
            "SELECT customer_id FROM crm_followups WHERE status='pending'").fetchall())
        count = 0
        for row in absent:
            if row[0] not in existing:
                conn.execute("INSERT INTO crm_followups (customer_id, type, scheduled_date, reason) VALUES (?,?,?,?)",
                            (row[0], 'absence', scheduled, 'Client absent > 60 jours'))
                count += 1
        conn.commit()
    flash(f"{count} suivis générés automatiquement", "success")
    return redirect("/crm_followups")



# ─── 6. Referral System ───
@customers_bp.route("/referrals")
@login_required
def referrals():
    with get_db() as conn:
        refs = conn.execute("""SELECT r.*, c.name as referrer_name, c.phone as referrer_phone
            FROM referrals r JOIN customers c ON r.referrer_id=c.id ORDER BY r.created_at DESC""").fetchall()
        customers = conn.execute("SELECT id, name, phone FROM customers ORDER BY name").fetchall()
    return render_template("referrals.html", referrals=refs, customers=customers)



@customers_bp.route("/referral/add", methods=["POST"])
@login_required
def referral_add():
    referrer_id = request.form.get("referrer_id", type=int)
    referred_name = request.form.get("referred_name", "").strip()
    referred_phone = request.form.get("referred_phone", "").strip()
    reward_type = request.form.get("reward_type", "free_wash")
    if referrer_id and referred_name and referred_phone:
        with get_db() as conn:
            conn.execute("INSERT INTO referrals (referrer_id, referred_name, referred_phone, reward_type) VALUES (?,?,?,?)",
                        (referrer_id, referred_name, referred_phone, reward_type))
            conn.commit()
        flash("Parrainage enregistré !", "success")
    return redirect("/referrals")



@customers_bp.route("/referral/convert/<int:rid>", methods=["POST"])
@login_required
def referral_convert(rid):
    with get_db() as conn:
        ref = conn.execute("SELECT * FROM referrals WHERE id=?", (rid,)).fetchone()
        if ref:
            # Create customer from referral
            conn.execute("INSERT INTO customers (name, phone, referred_by) VALUES (?,?,?)",
                        (ref[2], ref[3], ref[1]))
            new_cid = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
            conn.execute("UPDATE referrals SET status='converted', converted_customer_id=? WHERE id=?", (new_cid, rid))
            # Reward the referrer
            try:
                conn.execute("INSERT INTO reward_points (customer_id, points, total_earned) VALUES (?, 50, 50) ON CONFLICT(customer_id) DO UPDATE SET points=points+50, total_earned=total_earned+50", (ref[1],))
                conn.execute("INSERT INTO reward_history (customer_id, points, type, description) VALUES (?,50,'earn','Bonus parrainage')", (ref[1],))
            except Exception:
                pass
            conn.commit()
            flash(f"Client créé + 50 points offerts au parrain !", "success")
    return redirect("/referrals")



# ─── 7. Fleet / Company Accounts ───
@customers_bp.route("/fleet_companies")
@login_required
def fleet_companies():
    with get_db() as conn:
        companies = conn.execute("SELECT * FROM fleet_companies ORDER BY name").fetchall()
        # Count vehicles per company
        vehicle_counts = {}
        for row in conn.execute("SELECT company_id, COUNT(*) FROM fleet_vehicles GROUP BY company_id").fetchall():
            vehicle_counts[row[0]] = row[1]
    return render_template("fleet_companies.html", companies=companies, vehicle_counts=vehicle_counts)



@customers_bp.route("/fleet_company/add", methods=["POST"])
@login_required
def fleet_company_add():
    name = request.form.get("name", "").strip()
    contact = request.form.get("contact_person", "").strip()
    phone = request.form.get("phone", "").strip()
    email = request.form.get("email", "").strip()
    address = request.form.get("address", "").strip()
    contract_start = request.form.get("contract_start", "")
    contract_end = request.form.get("contract_end", "")
    discount = request.form.get("discount_percent", 0, type=float)
    payment_terms = request.form.get("payment_terms", "monthly")
    notes = request.form.get("notes", "").strip()
    if name:
        with get_db() as conn:
            conn.execute("""INSERT INTO fleet_companies (name, contact_person, phone, email, address,
                contract_start, contract_end, discount_percent, payment_terms, notes)
                VALUES (?,?,?,?,?,?,?,?,?,?)""",
                (name, contact, phone, email, address, contract_start, contract_end, discount, payment_terms, notes))
            conn.commit()
        flash("Entreprise ajoutée !", "success")
    return redirect("/fleet_companies")



@customers_bp.route("/fleet_company/<int:cid>")
@login_required
def fleet_company_detail(cid):
    with get_db() as conn:
        company = conn.execute("SELECT * FROM fleet_companies WHERE id=?", (cid,)).fetchone()
        if not company:
            flash("Entreprise non trouvée", "error")
            return redirect("/fleet_companies")
        vehicles = conn.execute("""SELECT fv.id, ca.id as car_id, ca.brand, ca.model, ca.plate, c.name as owner
            FROM fleet_vehicles fv JOIN cars ca ON fv.car_id=ca.id
            JOIN customers c ON ca.customer_id=c.id WHERE fv.company_id=?""", (cid,)).fetchall()
        all_cars = conn.execute("SELECT ca.id, ca.brand, ca.model, ca.plate, c.name FROM cars ca JOIN customers c ON ca.customer_id=c.id ORDER BY c.name").fetchall()
        # Invoice summary for company vehicles
        car_ids = [v[1] for v in vehicles]
        total_spent = 0
        if car_ids:
            placeholders = ','.join('?' * len(car_ids))
            total_spent = conn.execute(f"""SELECT COALESCE(SUM(i.amount),0) FROM invoices i
                JOIN appointments a ON i.appointment_id=a.id
                WHERE a.car_id IN ({placeholders}) AND i.status='paid'""", car_ids).fetchone()[0]
    return render_template("fleet_company_detail.html", company=company, vehicles=vehicles,
                          all_cars=all_cars, total_spent=total_spent)



@customers_bp.route("/fleet_vehicle/add/<int:cid>", methods=["POST"])
@login_required
def fleet_vehicle_add(cid):
    car_id = request.form.get("car_id", type=int)
    if car_id:
        with get_db() as conn:
            try:
                conn.execute("INSERT INTO fleet_vehicles (company_id, car_id) VALUES (?,?)", (cid, car_id))
                conn.commit()
                flash("Véhicule ajouté à la flotte !", "success")
            except Exception:
                flash("Ce véhicule est déjà dans cette flotte", "error")
    return redirect(f"/fleet_company/{cid}")



@customers_bp.route("/fleet_vehicle/remove/<int:fvid>/<int:cid>", methods=["POST"])
@login_required
def fleet_vehicle_remove(fvid, cid):
    with get_db() as conn:
        conn.execute("DELETE FROM fleet_vehicles WHERE id=?", (fvid,))
        conn.commit()
    flash("Véhicule retiré de la flotte", "success")
    return redirect(f"/fleet_company/{cid}")



@customers_bp.route("/rfm_analysis")
@login_required
@admin_required
def rfm_analysis():
    from datetime import date, timedelta
    today = date.today()
    with get_db() as conn:
        # Calculate RFM for all customers
        customers = conn.execute("""SELECT c.id, c.name, c.phone,
            MAX(a.date) as last_visit,
            COUNT(DISTINCT a.id) as frequency,
            COALESCE(SUM(i.amount),0) as monetary
            FROM customers c
            LEFT JOIN cars ca ON ca.customer_id=c.id
            LEFT JOIN appointments a ON a.car_id=ca.id AND a.status='completed'
            LEFT JOIN invoices i ON i.appointment_id=a.id AND i.status='paid'
            GROUP BY c.id""").fetchall()

        segments = {}
        all_data = []
        for c in customers:
            cid, name, phone = c[0], c[1], c[2]
            last_visit = c[3]
            freq = c[4] or 0
            monetary = c[5] or 0

            # Recency in days
            if last_visit:
                try:
                    lv = date.fromisoformat(last_visit)
                    recency_days = (today - lv).days
                except (ValueError, TypeError):
                    recency_days = 999
            else:
                recency_days = 999

            # Score 1-5
            r = 5 if recency_days < 30 else 4 if recency_days < 60 else 3 if recency_days < 90 else 2 if recency_days < 180 else 1
            f = min(5, max(1, freq))
            m_score = 5 if monetary > 1000 else 4 if monetary > 500 else 3 if monetary > 200 else 2 if monetary > 50 else 1

            fm = max(f, m_score)
            segment = RFM_SEGMENTS.get((r, fm), 'need_attention')

            # Save to DB
            conn.execute("""INSERT INTO rfm_segments (customer_id, recency_score, frequency_score, monetary_score, rfm_score, segment)
                VALUES (?,?,?,?,?,?) ON CONFLICT(customer_id) DO UPDATE SET
                recency_score=?, frequency_score=?, monetary_score=?, rfm_score=?, segment=?, last_calculated=CURRENT_TIMESTAMP""",
                (cid, r, f, m_score, r*100+f*10+m_score, segment, r, f, m_score, r*100+f*10+m_score, segment))
            conn.execute("UPDATE customers SET rfm_segment=? WHERE id=?", (segment, cid))

            segments[segment] = segments.get(segment, 0) + 1
            all_data.append({'id': cid, 'name': name, 'phone': phone, 'segment': segment,
                            'recency': recency_days, 'frequency': freq, 'monetary': monetary,
                            'r': r, 'f': f, 'm': m_score})
        conn.commit()

    return render_template("rfm_analysis.html", data=all_data, segments=segments, labels=RFM_LABELS)



# ─── 3. Timeline Client Unifiée ───

@customers_bp.route("/customer_timeline/<int:cid>")
@login_required
def customer_timeline(cid):
    with get_db() as conn:
        customer = conn.execute("SELECT * FROM customers WHERE id=?", (cid,)).fetchone()
        if not customer:
            flash("Client introuvable", "danger")
            return redirect("/customers")
        events = []
        # Appointments
        for a in conn.execute("SELECT * FROM appointments WHERE car_id IN (SELECT id FROM cars WHERE customer_id=?) ORDER BY date DESC", (cid,)).fetchall():
            events.append({'type': 'appointment', 'icon': '📅', 'title': f"RDV — {a['service']}", 
                'detail': f"Statut: {a['status']}", 'date': a['date'], 'ref_id': a['id']})
        # Invoices
        for i in conn.execute("SELECT * FROM invoices WHERE appointment_id IN (SELECT id FROM appointments WHERE car_id IN (SELECT id FROM cars WHERE customer_id=?))", (cid,)).fetchall():
            events.append({'type': 'invoice', 'icon': '📄', 'title': f"Facture #{i['id']} — {i['amount']} DT",
                'detail': f"Statut: {i['status']}", 'date': i.get('created_at', ''), 'ref_id': i['id']})
        # Communications
        for c in conn.execute("SELECT * FROM communication_log WHERE customer_id=? ORDER BY created_at DESC", (cid,)).fetchall():
            events.append({'type': 'communication', 'icon': '💬', 'title': f"{c['comm_type']} — {c['subject']}",
                'detail': c.get('message', '')[:100], 'date': c['created_at'], 'ref_id': c['id']})
        # CRM Follow-ups
        for f in conn.execute("SELECT * FROM crm_followups WHERE customer_id=?", (cid,)).fetchall():
            events.append({'type': 'followup', 'icon': '🔄', 'title': f"Suivi — {f['action_type']}",
                'detail': f['notes'][:100] if f['notes'] else '', 'date': f['scheduled_date'], 'ref_id': f['id']})
        # Ratings
        for r in conn.execute("SELECT * FROM ratings WHERE customer_id=?", (cid,)).fetchall():
            events.append({'type': 'rating', 'icon': '⭐', 'title': f"Évaluation — {r['score']}/5",
                'detail': r.get('comment', ''), 'date': r.get('created_at', ''), 'ref_id': r['id']})
        # Insurance claims
        for ic in conn.execute("SELECT * FROM insurance_claims WHERE customer_id=?", (cid,)).fetchall():
            events.append({'type': 'insurance', 'icon': '🏥', 'title': f"Dossier assurance #{ic['claim_number']}",
                'detail': f"Statut: {ic['status']} — {ic['estimated_cost']} DT", 'date': ic['created_at'], 'ref_id': ic['id']})
        # Sort by date desc
        events.sort(key=lambda x: x.get('date', '') or '', reverse=True)
        cars = conn.execute("SELECT * FROM cars WHERE customer_id=?", (cid,)).fetchall()
    return render_template("customer_timeline.html", customer=customer, events=events, cars=cars)



@customers_bp.route("/vip_program")
@login_required
def vip_program():
    with get_db() as conn:
        levels = conn.execute("SELECT * FROM vip_levels ORDER BY min_spend ASC").fetchall()
        if not levels:
            for lv in DEFAULT_VIP_LEVELS:
                conn.execute("""INSERT INTO vip_levels (name, min_spend, discount_percent, perks, color, icon, sort_order)
                    VALUES (?,?,?,?,?,?,?)""", (lv['name'], lv['min_spend'], lv['discount'], lv['perks'], lv['color'], lv['icon'], DEFAULT_VIP_LEVELS.index(lv)))
            conn.commit()
            levels = conn.execute("SELECT * FROM vip_levels ORDER BY min_spend ASC").fetchall()
        # Calculate customer VIP levels
        customers = conn.execute("""SELECT cu.*, COALESCE(SUM(i.amount), 0) as total_spent 
            FROM customers cu 
            LEFT JOIN cars c ON c.customer_id = cu.id 
            LEFT JOIN appointments a ON a.car_id = c.id 
            LEFT JOIN invoices i ON i.appointment_id = a.id AND i.status = 'paid' 
            GROUP BY cu.id ORDER BY total_spent DESC""").fetchall()
        # Assign VIP levels
        vip_customers = []
        level_counts = {l['name']: 0 for l in levels}
        for cu in customers:
            spent = cu['total_spent']
            assigned = levels[0]
            for lv in levels:
                if spent >= lv['min_spend']:
                    assigned = lv
            level_counts[assigned['name']] = level_counts.get(assigned['name'], 0) + 1
            vip_customers.append({'customer': cu, 'level': assigned, 'spent': spent})
            # Update customer vip_level
            conn.execute("UPDATE customers SET vip_level=?, total_spent=? WHERE id=?", 
                        (assigned['name'], spent, cu['id']))
        conn.commit()
    return render_template("vip_program.html", levels=levels, vip_customers=vip_customers, level_counts=level_counts)



@customers_bp.route("/vip_level/edit", methods=["POST"])
@login_required
@admin_required
def edit_vip_level():
    lid = request.form.get("level_id", type=int)
    min_spend = request.form.get("min_spend", 0, type=float)
    discount = request.form.get("discount_percent", 0, type=float)
    perks = request.form.get("perks", "").strip()
    if lid:
        with get_db() as conn:
            conn.execute("UPDATE vip_levels SET min_spend=?, discount_percent=?, perks=? WHERE id=?",
                        (min_spend, discount, perks, lid))
            conn.commit()
        flash("Niveau mis à jour", "success")
    return redirect("/vip_program")



# ─── 3. Prédiction Churn Client ───

@customers_bp.route("/churn_prediction")
@login_required
@admin_required
def churn_prediction():
    from datetime import date, timedelta
    with get_db() as conn:
        today = date.today()
        customers = conn.execute("""SELECT cu.*, 
            MAX(a.date) as last_visit_date,
            COUNT(a.id) as visit_count,
            COALESCE(SUM(i.amount),0) as lifetime_value
            FROM customers cu 
            LEFT JOIN cars c ON c.customer_id=cu.id
            LEFT JOIN appointments a ON a.car_id=c.id AND a.status='completed'
            LEFT JOIN invoices i ON i.appointment_id=a.id AND i.status='paid'
            GROUP BY cu.id HAVING visit_count > 0
            ORDER BY last_visit_date ASC""").fetchall()
        predictions = []
        for cu in customers:
            last_visit = cu['last_visit_date'] or ''
            if not last_visit:
                continue
            try:
                last_dt = date.fromisoformat(last_visit)
            except (ValueError, TypeError, AttributeError):
                continue
            days_since = (today - last_dt).days
            visits = cu['visit_count']
            # Calculate average interval
            all_dates = conn.execute("""SELECT DISTINCT a.date FROM appointments a 
                JOIN cars c ON a.car_id=c.id WHERE c.customer_id=? AND a.status='completed' 
                ORDER BY a.date""", (cu['id'],)).fetchall()
            if len(all_dates) > 1:
                intervals = []
                for j in range(1, len(all_dates)):
                    try:
                        d1 = date.fromisoformat(all_dates[j-1]['date'])
                        d2 = date.fromisoformat(all_dates[j]['date'])
                        intervals.append((d2-d1).days)
                    except (ValueError, TypeError, AttributeError):
                        pass
                avg_interval = sum(intervals)/len(intervals) if intervals else 90
            else:
                avg_interval = 90
            # Risk score: higher = more likely to churn
            if avg_interval > 0:
                ratio = days_since / avg_interval
            else:
                ratio = days_since / 90
            if ratio >= 3:
                risk_score = min(100, 60 + ratio * 5)
                risk_level = 'critical'
            elif ratio >= 2:
                risk_score = 40 + ratio * 10
                risk_level = 'high'
            elif ratio >= 1.5:
                risk_score = 30 + ratio * 5
                risk_level = 'medium'
            else:
                risk_score = ratio * 20
                risk_level = 'low'
            risk_score = min(100, max(0, risk_score))
            predicted_churn = (last_dt + timedelta(days=int(avg_interval * 2.5))).isoformat()
            predictions.append({
                'customer': cu, 'days_since': days_since, 'visits': visits,
                'avg_interval': avg_interval, 'risk_score': risk_score,
                'risk_level': risk_level, 'predicted_churn': predicted_churn,
                'lifetime_value': cu['lifetime_value']
            })
            # Save
            existing = conn.execute("SELECT id FROM churn_predictions WHERE customer_id=?", (cu['id'],)).fetchone()
            if existing:
                conn.execute("""UPDATE churn_predictions SET risk_score=?, risk_level=?, 
                    days_since_last_visit=?, avg_visit_interval=?, predicted_churn_date=? WHERE customer_id=?""",
                    (risk_score, risk_level, days_since, avg_interval, predicted_churn, cu['id']))
            else:
                conn.execute("""INSERT INTO churn_predictions 
                    (customer_id, risk_score, risk_level, days_since_last_visit, avg_visit_interval, predicted_churn_date)
                    VALUES (?,?,?,?,?,?)""",
                    (cu['id'], risk_score, risk_level, days_since, avg_interval, predicted_churn))
            conn.execute("UPDATE customers SET churn_risk=?, last_churn_check=? WHERE id=?",
                        (risk_level, today.isoformat(), cu['id']))
        conn.commit()
        predictions.sort(key=lambda x: x['risk_score'], reverse=True)
    return render_template("churn_prediction.html", predictions=predictions)



# ─── 4. NPS & Satisfaction ───

@customers_bp.route("/nps_dashboard")
@login_required
def nps_dashboard():
    with get_db() as conn:
        surveys = conn.execute("""SELECT n.*, c.name as customer_name, c.phone
            FROM nps_surveys n LEFT JOIN customers c ON n.customer_id=c.id
            ORDER BY n.created_at DESC LIMIT 100""").fetchall()
        total = len(surveys)
        if total > 0:
            promoters = sum(1 for s in surveys if s['score'] >= 9)
            passives = sum(1 for s in surveys if 7 <= s['score'] <= 8)
            detractors = sum(1 for s in surveys if s['score'] <= 6)
            nps = int(((promoters - detractors) / total) * 100)
            avg_score = sum(s['score'] for s in surveys) / total
        else:
            promoters = passives = detractors = 0
            nps = 0
            avg_score = 0
        monthly = conn.execute("""SELECT strftime('%%Y-%%m', created_at) as month,
            AVG(score) as avg_score, COUNT(*) as count FROM nps_surveys
            GROUP BY month ORDER BY month DESC LIMIT 6""").fetchall()
        alerts = conn.execute("""SELECT n.*, c.name as customer_name, c.phone
            FROM nps_surveys n LEFT JOIN customers c ON n.customer_id=c.id
            WHERE n.score <= 6 AND n.follow_up_status='none'
            ORDER BY n.created_at DESC""").fetchall()
    return render_template("nps_dashboard.html", surveys=surveys, nps=nps, avg_score=round(avg_score, 1),
        promoters=promoters, passives=passives, detractors=detractors, total=total,
        monthly=monthly, alerts=alerts)



@customers_bp.route("/nps_survey/add", methods=["POST"])
@login_required
def nps_survey_add():
    with get_db() as conn:
        customer_id = request.form.get("customer_id", 0, type=int)
        score = request.form.get("score", 0, type=int)
        feedback = request.form.get("feedback", "")
        appointment_id = request.form.get("appointment_id", 0, type=int)
        category = 'promoter' if score >= 9 else 'passive' if score >= 7 else 'detractor'
        conn.execute("""INSERT INTO nps_surveys (customer_id, appointment_id, score, category, feedback)
            VALUES (?,?,?,?,?)""", (customer_id, appointment_id, score, category, feedback))
        conn.execute("UPDATE customers SET nps_score=? WHERE id=?", (score, customer_id))
        conn.commit()
    flash("Enquête NPS enregistrée", "success")
    return redirect("/nps_dashboard")



@customers_bp.route("/nps_followup/<int:survey_id>", methods=["POST"])
@login_required
def nps_followup(survey_id):
    with get_db() as conn:
        conn.execute("UPDATE nps_surveys SET follow_up_status=?, follow_up_notes=? WHERE id=?",
            (request.form.get("status", "contacted"), request.form.get("notes", ""), survey_id))
        conn.commit()
    flash("Suivi mis à jour", "success")
    return redirect("/nps_dashboard")



# ─── 5. Portefeuille Client (Wallet) ───

@customers_bp.route("/wallet/<int:customer_id>")
@login_required
def wallet_view(customer_id):
    with get_db() as conn:
        customer = conn.execute("SELECT * FROM customers WHERE id=?", (customer_id,)).fetchone()
        if not customer:
            flash("Client non trouvé", "danger")
            return redirect("/customers")
        transactions = conn.execute("""SELECT * FROM wallet_transactions
            WHERE customer_id=? ORDER BY created_at DESC LIMIT 50""", (customer_id,)).fetchall()
        balance = customer['wallet_balance'] or 0 if customer['wallet_balance'] else 0
    return render_template("wallet.html", customer=customer, transactions=transactions, balance=balance)



@customers_bp.route("/wallet/topup", methods=["POST"])
@login_required
def wallet_topup():
    customer_id = request.form.get("customer_id", 0, type=int)
    amount = request.form.get("amount", 0, type=float)
    description = request.form.get("description", "Recharge manuelle")
    if amount <= 0:
        flash("Montant invalide", "danger")
        return redirect(f"/wallet/{customer_id}")
    with get_db() as conn:
        current = conn.execute("SELECT wallet_balance FROM customers WHERE id=?", (customer_id,)).fetchone()
        if not current:
            flash("Client non trouvé", "danger")
            return redirect("/customers")
        new_balance = (current['wallet_balance'] or 0) + amount
        conn.execute("UPDATE customers SET wallet_balance=? WHERE id=?", (new_balance, customer_id))
        conn.execute("""INSERT INTO wallet_transactions (customer_id, transaction_type, amount, balance_after, description, created_by)
            VALUES (?,?,?,?,?,?)""", (customer_id, 'topup', amount, new_balance, description, session.get('username', 'admin')))
        conn.commit()
    flash(f"+{amount} DH ajouté au portefeuille", "success")
    return redirect(f"/wallet/{customer_id}")



@customers_bp.route("/wallet/debit", methods=["POST"])
@login_required
def wallet_debit():
    customer_id = request.form.get("customer_id", 0, type=int)
    amount = request.form.get("amount", 0, type=float)
    description = request.form.get("description", "Paiement service")
    ref_type = request.form.get("reference_type", "")
    ref_id = request.form.get("reference_id", 0, type=int)
    with get_db() as conn:
        current = conn.execute("SELECT wallet_balance FROM customers WHERE id=?", (customer_id,)).fetchone()
        if not current:
            flash("Client non trouvé", "danger")
            return redirect("/customers")
        bal = current['wallet_balance'] or 0
        if amount > bal:
            flash("Solde insuffisant", "danger")
            return redirect(f"/wallet/{customer_id}")
        new_balance = bal - amount
        conn.execute("UPDATE customers SET wallet_balance=? WHERE id=?", (new_balance, customer_id))
        conn.execute("""INSERT INTO wallet_transactions (customer_id, transaction_type, amount, balance_after, description, reference_type, reference_id, created_by)
            VALUES (?,?,?,?,?,?,?,?)""", (customer_id, 'debit', -amount, new_balance, description, ref_type, ref_id, session.get('username', 'admin')))
        conn.commit()
    flash(f"-{amount} DH débité du portefeuille", "success")
    return redirect(f"/wallet/{customer_id}")



# ─── 10. Fidélité Gamifiée ───

@customers_bp.route("/loyalty_gamified")
@login_required
def loyalty_gamified():
    with get_db() as conn:
        challenges = conn.execute("SELECT * FROM loyalty_challenges ORDER BY created_at DESC").fetchall()
        levels = conn.execute("""SELECT loyalty_level, COUNT(*) as count FROM customers
            WHERE loyalty_level != '' GROUP BY loyalty_level ORDER BY
            CASE loyalty_level WHEN 'platinum' THEN 1 WHEN 'gold' THEN 2
            WHEN 'silver' THEN 3 ELSE 4 END""").fetchall()
        top_loyal = conn.execute("""SELECT name, phone, loyalty_level, loyalty_points_total, wallet_balance
            FROM customers WHERE loyalty_points_total > 0
            ORDER BY loyalty_points_total DESC LIMIT 20""").fetchall()
    return render_template("loyalty_gamified.html", challenges=challenges, levels=levels, top_loyal=top_loyal)



@customers_bp.route("/loyalty_challenge/add", methods=["POST"])
@login_required
def loyalty_challenge_add():
    with get_db() as conn:
        conn.execute("""INSERT INTO loyalty_challenges
            (title, description, challenge_type, target_value, reward_points, reward_description, start_date, end_date, vehicle_types)
            VALUES (?,?,?,?,?,?,?,?,?)""",
            (request.form["title"], request.form.get("description", ""),
             request.form["challenge_type"], request.form.get("target_value", 0, type=float),
             request.form.get("reward_points", 0, type=int), request.form.get("reward_description", ""),
             request.form["start_date"], request.form["end_date"],
             request.form.get("vehicle_types", "all")))
        conn.commit()
    flash("Challenge créé !", "success")
    return redirect("/loyalty_gamified")



@customers_bp.route("/loyalty_level_update")
@login_required
def loyalty_level_update():
    with get_db() as conn:
        customers = conn.execute("SELECT id, loyalty_points_total FROM customers").fetchall()
        updated = 0
        for c in customers:
            pts = c['loyalty_points_total'] or 0
            if pts >= 5000:
                level = 'platinum'
            elif pts >= 2000:
                level = 'gold'
            elif pts >= 500:
                level = 'silver'
            else:
                level = 'bronze'
            conn.execute("UPDATE customers SET loyalty_level=? WHERE id=?", (level, c['id']))
            updated += 1
        conn.commit()
    flash(f"{updated} niveaux de fidélité mis à jour", "success")
    return redirect("/loyalty_gamified")


